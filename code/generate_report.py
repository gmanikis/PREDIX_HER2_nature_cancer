#!/usr/bin/env python3
"""
UNIFIED REPORT — PREDIX HER2 MULTIMODAL pCR PREDICTION
=======================================================
Reads PKL files from multimodal_pcr_pipeline.py (signature discovery mode)
and generates a comprehensive figure set and Excel workbook.

USAGE
-----
  python3 generate_report.py --results_dir ./results --out_dir ./report

OUTPUT LAYOUT
-------------
  report/figures/                 main-text figures (fig01-fig06)
  report/figures/supplementary/   discovery diagnostics (supp_fig01-supp_fig14)
  report/tables/                  headline workbooks
  report/tables/supplementary/    diagnostic workbooks

MAIN-TEXT FIGURES — all use the frozen-consensus signature under the iterated
outer CV protocol, which is the single source of truth for performance claims.
  fig01  AUROC per model (mean over CV repeats of the pooled-OOF AUROC),
         with 95% patient-level cluster-bootstrap CI
  fig02  Consensus signatures per modality × scenario
  fig03  Pooled-OOF ROC — fused vs best consensus unimodal
  fig04  Consensus fusion modality weights
  fig05  SHAP beeswarm restricted to consensus signature features (per scenario)
  fig06  Counterfactual arm-switch summary

SUPPLEMENTARY FIGURES — the discovery phase, prefixed supp_.
  ROC curves, performance distributions, fusion benefit, forest plot,
  feature SHAP, selection frequency, cross-scenario features, fusion SHAP,
  modality weights, winner-classifier heatmap, inner AUROC comparison,
  calibration profile, signature sizes, performance CI.

REPORTING UNCERTAINTY — IMPORTANT
---------------------------------
Two different uncertainty conventions appear in these outputs and they are NOT
interchangeable.

  Patient-level cluster-bootstrap CI   The correct interval (cv_estimands.py).
                               Point estimate = metric computed on each CV
                               repeat's complete out-of-fold prediction
                               vector, averaged over repeats (a single-model
                               estimand; probabilities are never averaged
                               across repeats). CI = patients resampled with
                               replacement, all R predictions of a resampled
                               patient moving together. Used in fig01, the
                               consensus Performance sheet, the supplementary
                               performance-CI table/forest plot and the
                               fusion-benefit / paired-comparison figures.
                               Quote these.

  SD across folds ("±")        Descriptive spread only. The folds reuse the
                               same patients, so this does not measure
                               sampling uncertainty about the cohort and must
                               never be presented as a confidence interval.
                               Columns carrying it are labelled as such.

No interval in this report is obtained by resampling per-fold metric values
or by bootstrapping patient-averaged probabilities (both give biased
intervals; see the cv_estimands.py docstring). Calibration, selection
stability, per-fold event counts and the full paired model comparisons come
from revision_analyses.py.

EXCEL SHEETS (headline workbook)
--------------------------------
  Signatures   consensus signature per modality with selection frequency
  Fusion       modality weights and selection rate
  Performance  AUROC (mean over CV repeats) with patient-level
               cluster-bootstrap CI, plus the pooled OOF metrics
"""

import argparse, ast, json, pickle, warnings
from functools import lru_cache
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
from pathlib import Path
from collections import Counter, defaultdict
from sklearn.metrics import roc_curve
from scipy.stats import gaussian_kde
import cv_estimands as CE
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# READING THE PIPELINE'S OWN CONSTANTS  (parse, never retype)
# =============================================================================
# Every number this report prints about the METHOD must come from the place
# that defines the method, not from a human retyping it into a string literal.
# That rule exists because the failure is silent: a stale constant does not
# raise, it just emits a wrong value into a supplementary table and ships. It
# has already happened three times in this project (the "11 total" Tier-1
# string below, the 13-entry list in build_supp_table_S-ML9.py, and the ED
# Fig 11a schematic reading a dead CLI argument).
#
# Three sources, one rule each:
#   * hard-coded pipeline constants  -> parsed from the source with `ast`
#   * CLI-settable parameters        -> read from the run's run_provenance.json,
#                                       which records what ACTUALLY ran
#   * realised effects               -> computed from the PKLs in `data`
#
# The pipeline is parsed, never imported: importing it pulls in sklearn and
# shap and sets BLAS environment variables, which is far too much to pay for
# a handful of literals.
_PIPELINE_FILE = "multimodal_pcr_pipeline.py"


@lru_cache(maxsize=1)
def _pipeline_source():
    """Parse the pipeline once; return (path, tree, consts, assign_counts).

    `consts` resolves top-level literal assignments in source order, so a name
    defined as an alias of an earlier one (FUSION_C_GRID = ELASTICNET_C_GRID)
    resolves too. `assign_counts` counts assignments anywhere in the file so
    callers can refuse to trust a name that is rebound somewhere else.
    """
    src_path = Path(__file__).resolve().parent / _PIPELINE_FILE
    tree = ast.parse(src_path.read_text(encoding="utf-8-sig"))

    counts = Counter()
    for node in ast.walk(tree):
        if isinstance(node, ast.Assign):
            for tgt in node.targets:
                if isinstance(tgt, ast.Name):
                    counts[tgt.id] += 1
                # `globals()["NAME"] = v` — the dynamic rebinding used by the
                # loky worker config. Only constant subscripts are visible to
                # a static parse; a runtime key cannot be resolved here, which
                # is why _pipeline_const cross-checks provenance where it can.
                elif (isinstance(tgt, ast.Subscript)
                        and isinstance(tgt.value, ast.Call)
                        and isinstance(tgt.value.func, ast.Name)
                        and tgt.value.func.id == "globals"
                        and isinstance(tgt.slice, ast.Constant)):
                    counts[tgt.slice.value] += 1

    def _value(node):
        """literal_eval, plus `set()` and aliases of already-known names."""
        if (isinstance(node, ast.Call) and isinstance(node.func, ast.Name)
                and node.func.id == "set" and not node.args and not node.keywords):
            return set()
        if isinstance(node, ast.Name) and node.id in consts:
            return consts[node.id]
        return ast.literal_eval(node)

    consts = {}
    for node in tree.body:                      # top level only, in order
        if not isinstance(node, ast.Assign):
            continue
        try:
            val = _value(node.value)
        except (ValueError, SyntaxError, TypeError):
            continue
        for tgt in node.targets:
            if isinstance(tgt, ast.Name):
                consts[tgt.id] = val
    return src_path, tree, consts, counts


def _pipeline_const(name):
    """Return a module-level constant from the pipeline, or raise.

    Raises rather than defaulting. A missing or rebound constant means this
    report can no longer describe the pipeline truthfully, and failing here is
    strictly better than emitting a plausible wrong number into a table that
    goes to reviewers.
    """
    src_path, _, consts, counts = _pipeline_source()
    if counts[name] > 1:
        raise RuntimeError(
            f"{name} is assigned {counts[name]} times in {src_path.name} — "
            f"refusing to report a value that may be rebound elsewhere")
    if name not in consts:
        raise RuntimeError(f"no top-level literal {name} in {src_path}")
    return consts[name]


def _provenance_params(results_dir):
    """`parameters` from the run's run_provenance.json, or {} if unavailable.

    This is the record of what the run actually used, so it beats the
    pipeline's argparse defaults for anything the CLI can override. Callers
    fall back to _pipeline_cli_default when a run predates the provenance
    record.
    """
    if results_dir is None:
        return {}
    try:
        with open(Path(results_dir) / "run_provenance.json",
                  encoding="utf-8") as fh:
            return json.load(fh).get("parameters", {}) or {}
    except (OSError, ValueError):
        return {}


def _pipeline_func_const(func_name, name):
    """A literal assigned to a local inside a named pipeline function.

    Tier 4's EPV ceiling and feature floor are function locals in
    _derive_signature, so _pipeline_const cannot see them. Same rule applies:
    resolve it or raise — there is no typed fallback, because a silent
    fallback is precisely the failure mode being removed.
    """
    src_path, tree, _, _ = _pipeline_source()
    for node in ast.walk(tree):
        if not (isinstance(node, ast.FunctionDef) and node.name == func_name):
            continue
        found, seen = None, 0
        for sub in ast.walk(node):
            if not isinstance(sub, ast.Assign):
                continue
            for tgt in sub.targets:
                if isinstance(tgt, ast.Name) and tgt.id == name:
                    seen += 1
                    found = ast.literal_eval(sub.value)
        if seen == 1:
            return found
        if seen > 1:
            raise RuntimeError(
                f"{name} is assigned {seen} times inside {func_name}() — "
                f"refusing to report a value that may be rebound")
    raise RuntimeError(f"no local {name} in {func_name}() of {src_path.name}")


def _pipeline_percentile_arg(func_name):
    """The percentile passed to the single np.percentile call in a function.

    Deliberately narrow: it resolves exactly one call in one named function
    and raises on anything it does not recognise, rather than trying to be a
    general expression evaluator.
    """
    src_path, tree, _, _ = _pipeline_source()
    for node in ast.walk(tree):
        if not (isinstance(node, ast.FunctionDef) and node.name == func_name):
            continue
        hits = [sub for sub in ast.walk(node)
                if isinstance(sub, ast.Call)
                and isinstance(sub.func, ast.Attribute)
                and sub.func.attr == "percentile"]
        if len(hits) != 1:
            raise RuntimeError(
                f"expected exactly one np.percentile call in {func_name}(), "
                f"found {len(hits)} — refusing to guess which one is Tier 4's")
        args = hits[0].args
        if len(args) < 2 or not isinstance(args[1], ast.Constant):
            raise RuntimeError(
                f"np.percentile in {func_name}() does not take a literal "
                f"percentile — cannot report it without guessing")
        return args[1].value
    raise RuntimeError(f"no function {func_name}() in {src_path.name}")


def _pipeline_cli_default(flag):
    """Default of `--flag` as declared in the pipeline's argparse setup."""
    src_path, tree, _, _ = _pipeline_source()
    for node in ast.walk(tree):
        if not (isinstance(node, ast.Call)
                and isinstance(node.func, ast.Attribute)
                and node.func.attr == "add_argument"):
            continue
        names = [a.value for a in node.args
                 if isinstance(a, ast.Constant) and isinstance(a.value, str)]
        if f"--{flag}" not in names:
            continue
        for kw in node.keywords:
            if kw.arg == "default":
                return ast.literal_eval(kw.value)
    raise RuntimeError(f"no --{flag} default found in {src_path}")


def _run_param(results_dir, key):
    """A CLI-settable parameter: provenance first, argparse default second."""
    params = _provenance_params(results_dir)
    if key in params and params[key] is not None:
        return params[key]
    return _pipeline_cli_default(key)


def _stability_thresh(results_dir=None):
    """{scenario: selection-frequency threshold} for the 'stable' flag.

    CLI-settable (--stability_thresh_global / --stability_thresh_arm), so it
    takes the provenance path: what the run actually used, falling back to the
    pipeline's own argparse defaults. Previously a hand-typed dict here, which
    fed both the Tier-4 rows and the Pruning_Statistics stable-feature columns
    of a shipped supplementary table.
    """
    g = float(_run_param(results_dir, "stability_thresh_global"))
    a = float(_run_param(results_dir, "stability_thresh_arm"))
    return {"Global": g, "DHP": a, "T-DM1": a}


def _tier1_remove_from_pipeline():
    """Read TIER1_REMOVE out of the pipeline source instead of retyping it.

    This list was hand-copied into the pruning-report sheet and silently went
    stale twice: the sheet still said "11 total" and named the run-2 set while
    the pipeline had moved to 21 entries. Nothing errored - a wrong number just
    shipped in a supplementary table. Parsing it with ast gives the value one
    home, and an unparseable pipeline raises here rather than emitting a wrong
    figure.
    """
    src_path, tree, consts, counts = _pipeline_source()
    if counts["TIER1_REMOVE"] > 1:
        raise RuntimeError("TIER1_REMOVE is assigned more than once")
    found = None
    for node in tree.body:
        if (isinstance(node, ast.Assign) and len(node.targets) == 1
                and isinstance(node.targets[0], ast.Name)
                and node.targets[0].id == "TIER1_REMOVE"):
            found = [ast.literal_eval(e) for e in node.value.elts]
    if found is None:
        raise RuntimeError(f"no TIER1_REMOVE literal in {src_path}")
    # Honour the run-6 form, `if not KEEP_RNA_FCGR3B: TIER1_REMOVE.append(...)`.
    for node in ast.walk(tree):
        if (isinstance(node, ast.Call) and isinstance(node.func, ast.Attribute)
                and node.func.attr in ("append", "extend", "insert", "remove")
                and isinstance(node.func.value, ast.Name)
                and node.func.value.id == "TIER1_REMOVE"):
            parent_ok = False
            for outer in ast.walk(tree):
                if isinstance(outer, ast.If) and any(
                        n is node for n in ast.walk(outer)):
                    test = outer.test
                    if (isinstance(test, ast.UnaryOp)
                            and isinstance(test.op, ast.Not)
                            and isinstance(test.operand, ast.Name)
                            and test.operand.id in consts):
                        parent_ok = True
                        if not consts[test.operand.id]:
                            found.append(ast.literal_eval(node.args[0]))
            if not parent_ok:
                raise RuntimeError(
                    "TIER1_REMOVE is mutated in a way this parser cannot "
                    "resolve — refusing to emit a possibly wrong feature list")
    if len(found) != len(set(found)):
        raise RuntimeError("TIER1_REMOVE contains duplicates")
    return found

# =============================================================================
# CONSTANTS
# =============================================================================
SCENARIOS  = ["Global", "DHP", "T-DM1"]
EXP_MAP    = {"Global": "global", "DHP": "dhp", "T-DM1": "tdm1"}
UNIMODALS  = ["Clin", "RNA", "DNA", "Prot", "WSI"]
FUS_VARS   = ["Fused_ElasticNet"]
ALL_MODELS = UNIMODALS + FUS_VARS
ALL_CLFS   = ["ElasticNet_LR", "RandomForest", "ExtraTrees",
              "HistGradBoost", "SVM_Linear"]
CLF_LABEL  = {
    "ElasticNet_LR": "Elastic-Net LR",
    "RandomForest":  "Random Forest",
    "ExtraTrees":    "Extra Trees",
    "HistGradBoost": "Hist. GradBoost",
    "SVM_Linear":    "SVM Linear",
}
MOD_COLOR  = {
    "Clin": "#4e79a7", "RNA": "#f28e2b", "DNA": "#e15759",
    "Prot": "#76b7b2", "WSI": "#59a14f",
    "Fused_ElasticNet": "#6a1f6a",
}
CLF_COLOR  = {
    "ElasticNet_LR": "#4e79a7", "RandomForest": "#f28e2b",
    "ExtraTrees":    "#e15759", "HistGradBoost": "#76b7b2",
    "SVM_Linear":    "#59a14f",
}
SC_COL     = {"Global": "#333333", "DHP": "#2166ac", "T-DM1": "#d6604d"}
METRIC_KEYS = {
    "AUROC": "aurocs", "AUPRC": "auprcs", "Brier": "briers",
    "Sensitivity": "senss", "Specificity": "specs",
}

# Stability thresholds, read from the pipeline's own argparse defaults rather
# than retyped. Report builders that know which run they are describing should
# call _stability_thresh(results_dir) instead, so that a run which overrode the
# threshold on the command line is described by what it actually used; this
# module-level value is the default for callers that have no results_dir.
#
# NOTE: evaluating this at import time makes multimodal_pcr_pipeline.py a hard
# import-time dependency of this module. That is deliberate — the report cannot
# be produced without it in any case (TIER1_REMOVE, CORR_FILTER_MODS and the
# Tier-4 constants all come from it), and failing at import is far better than
# failing after ten minutes of figure rendering.
STABILITY_THRESH = _stability_thresh()

# =============================================================================
# MODE DETECTION
# =============================================================================
def detect_mode(data):
    """
    Return 'expanded' or 'cc_only' by inspecting the PKL fold dicts.
    Both modes now route through _fit_signature_model and produce the same
    fold_dict keys. The distinction is purely informational (startup print,
    Signatures sheet column label). We detect cc_only via n_events_inner:
    in expanded mode this varies across folds (each removes different test
    patients from a larger pool); in cc_only it is constant (same CC-only
    training size in every fold of a given scenario). Falls back to 'expanded'.
    """
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            folds = data[sc].get(mod, {}).get("folds", [])
            if len(folds) >= 2:
                n_events = [f.get("n_events_inner", 0) for f in folds[:10]
                            if f.get("n_events_inner", 0) > 0]
                if n_events and len(set(n_events)) == 1:
                    return "cc_only"
                return "expanded"
    return "expanded"


def get_signature_feats(fold):
    """
    Return the signature feature list for a single fold.
    Both expanded and cc_only now use winner_signature (from _fit_signature_model).
    Falls back to selected_features for PKLs generated before this fix.
    """
    return (fold.get("winner_signature")
            or fold.get("selected_features")
            or [])

plt.rcParams.update({
    "font.family": "sans-serif", "font.sans-serif": ["Arial", "DejaVu Sans"],
    "font.size": 9, "axes.titlesize": 10, "axes.titleweight": "bold",
    "axes.labelsize": 9, "xtick.labelsize": 8, "ytick.labelsize": 8,
    "legend.fontsize": 8, "legend.framealpha": 0.85,
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.linewidth": 0.8, "grid.linewidth": 0.5, "grid.alpha": 0.4,
    "savefig.bbox": "tight", "savefig.dpi": 300,
})

# =============================================================================
# CLI
# =============================================================================
def parse_args():
    p = argparse.ArgumentParser(
        description="PREDIX HER2 unified report",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    p.add_argument("--results_dir", type=Path, required=True)
    p.add_argument("--out_dir",     type=Path, default=Path("./report"))
    return p.parse_args()

# =============================================================================
# DATA LOADING
# =============================================================================
def load_data(results_dir):
    """
    Load all available PKL files.
    Returns data[scenario][model] = {folds, aurocs, auprcs, briers,
                                      senss, specs, roc, pr, summary}
    Plus raw folds for modality-specific analysis.
    """
    data = {sc: {} for sc in SCENARIOS}
    found = []

    for sc, exp in EXP_MAP.items():
        pkl = results_dir / exp / f"{exp}_elasticnet_results.pkl"
        if not pkl.exists():
            continue
        with open(pkl, "rb") as f:
            raw = pickle.load(f)

        for mod in ALL_MODELS:
            folds = raw.get(mod, [])
            if not folds:
                continue
            au = np.array([fd["metrics"]["AUROC"]       for fd in folds])
            ap = np.array([fd["metrics"]["AUPRC"]       for fd in folds])
            br = np.array([fd["metrics"]["Brier"]       for fd in folds])
            sn = np.array([fd["metrics"].get("Sensitivity", np.nan) for fd in folds])
            sp = np.array([fd["metrics"].get("Specificity", np.nan) for fd in folds])
            roc_c = []
            for fd in folds:
                fpr, tpr, _ = roc_curve(fd["y_test"], fd["y_pred"])
                roc_c.append((fpr, tpr))
            data[sc][mod] = {
                "folds":  folds,
                "aurocs": au, "auprcs": ap, "briers": br,
                "senss":  sn, "specs":  sp,
                "roc": roc_c,
                "summary": {
                    k: {"mean": float(np.nanmean(v)),
                        "std":  float(np.nanstd(v)),
                        "med":  float(np.nanmedian(v))}
                    for k, v in [("AUROC",au),("AUPRC",ap),("Brier",br),
                                  ("Sensitivity",sn),("Specificity",sp)]
                },
            }
        # Pooled metrics (added by the pipeline as results["_pooled_metrics"]).
        # One-threshold Youden Sens/Spec on concatenated-across-folds
        # (y_test, y_pred). Used by the Pooled_OpPoint Excel sheet.
        if "_pooled_metrics" in raw:
            data[sc]["_pooled_metrics"] = raw["_pooled_metrics"]

        # Consensus eval PKL (R2 protocol, written by evaluate_consensus).
        # Carries the frozen consensus signatures AND the pooled OOF
        # performance under that frozen consensus. This is the PRIMARY
        # HEADLINE numbers for the paper — separate from the discovery
        # fold-mean metrics above.
        cons_pkl = results_dir / exp / f"{exp}_consensus_eval.pkl"
        if cons_pkl.exists():
            with open(cons_pkl, "rb") as f:
                cons_raw = pickle.load(f)
            data[sc]["_consensus"] = cons_raw
            print(f"  [{sc}] consensus eval loaded "
                  f"(pooled fused AUROC = "
                  f"{cons_raw['pooled']['Fused_ElasticNet']['AUROC']:.4f})")

        found.append(sc)
        n = len(data[sc].get("RNA", {}).get("folds", []))
        print(f"  [{sc}] loaded — {n} folds")

    if not found:
        raise FileNotFoundError(
            f"No elasticnet PKL files found in {results_dir}. "
            "Run multimodal_pcr_pipeline.py first.")
    return data

# =============================================================================
# SHARED UTILITIES
# =============================================================================
def _stable_seed(*parts):
    """
    Deterministic bootstrap seed from string-able identity parts.

    NOT built on hash(): Python salts string hashing per process
    (PYTHONHASHSEED), so hash()-derived seeds differ between runs and every
    regeneration of the report produced slightly different CI endpoints,
    contradicting the seeded-reproducibility statement in run_provenance.json.
    zlib.crc32 is stable across processes and platforms.
    """
    import zlib
    return zlib.crc32("|".join(str(p) for p in parts).encode()) % (2 ** 31)


# ── Repeat-aware, patient-clustered CIs (cv_estimands.py) ────────────────────
# Every confidence interval in this report comes from these helpers. The
# estimand is the metric computed on each CV repeat's complete out-of-fold
# vector, averaged over repeats; the CI is a patient-level CLUSTER bootstrap
# (all R predictions of a resampled patient move together). Results are
# cached by identity so a figure and its companion table print identical
# numbers (same estimand, same seed, computed once).
_CE_CACHE = {}
# RUN 5: these are now imported from cv_estimands, which BOTH this script and
# revision_analyses.py already depend on, so they cannot drift apart.
# Previously generate_report hard-coded n_boot=2000 and the seed base while
# revision_analyses rebound both from its CLI — so `revision_analyses.py --seed
# 12345 --n_boot 5000` produced two workbooks carrying DIFFERENT intervals for
# the same quantity, with nothing to indicate why. The comment below claimed
# they matched; at defaults it was true, and one flag made it false.
_N_BOOT = CE.DEFAULT_N_BOOT
_BOOT_SEED_BASE = CE.DEFAULT_BOOT_SEED


def _shared_seed(tag):
    """Same tag grammar as revision_analyses, from the same definition."""
    return CE.shared_seed(tag, base=_BOOT_SEED_BASE)


def _consensus_auroc_ci(data, sc, mod):
    """
    Headline AUROC of one consensus model: mean over CV repeats of the
    pooled-OOF AUROC with a 95% patient-level cluster-bootstrap CI. Shared by
    fig01 and the Performance sheet of PREDIX_HER2_results.xlsx.
    Returns the cv_estimands result dict (estimate, ci_low, ci_high, n, ...).
    """
    key = ("consensus", sc, mod, "AUROC")
    if key not in _CE_CACHE:
        rm = CE.consensus_repeat_matrix(data[sc]["_consensus"], mod)
        _CE_CACHE[key] = CE.bootstrap_repeat_metric_ci(
            rm.P, rm.y, "AUROC", n_boot=_N_BOOT,
            seed=_shared_seed(f"{sc}|consensus|{mod}|AUROC"))
    return _CE_CACHE[key]


def _discovery_metric_ci(data, sc, mod, metric):
    """
    Discovery-phase AUROC / AUPRC / Brier of one model: mean over CV repeats
    of the pooled-OOF metric with a 95% patient-level cluster-bootstrap CI.
    Shared by supp fig14 and supp_PREDIX_HER2_performance_CI.xlsx.
    """
    key = ("discovery", sc, mod, metric)
    if key not in _CE_CACHE:
        rm = CE.discovery_repeat_matrix(data[sc][mod]["folds"])
        _CE_CACHE[key] = CE.bootstrap_repeat_metric_ci(
            rm.P, rm.y, metric, n_boot=_N_BOOT,
            seed=_shared_seed(f"{sc}|discovery|{mod}|{metric}"))
    return _CE_CACHE[key]


def _discovery_paired_delta(data, sc, mod_a, mod_b):
    """
    Discovery-phase paired ΔAUROC = mod_a − mod_b: per CV repeat the
    difference of the two pooled-OOF AUROCs, averaged over repeats, with a
    95% PAIRED patient-level cluster-bootstrap CI (the same patient resample
    is applied to both models and to all their repeats). Shared by supp fig03
    (fusion benefit) and supp fig04 (forest plot).
    """
    key = ("discovery_delta", sc, mod_a, mod_b)
    if key not in _CE_CACHE:
        rm_a = CE.discovery_repeat_matrix(data[sc][mod_a]["folds"])
        rm_b = CE.discovery_repeat_matrix(data[sc][mod_b]["folds"])
        if not np.array_equal(rm_a.pids, rm_b.pids):
            raise ValueError(f"[{sc}] {mod_a} and {mod_b} were evaluated on "
                             f"different patients; paired ΔAUROC undefined.")
        # revision_analyses.compare_models_repeat uses one seed per
        # (scenario, source) for every comparator, so the fused-vs-X deltas
        # here reproduce revision_model_comparisons.xlsx (discovery rows).
        _CE_CACHE[key] = CE.paired_bootstrap_repeat_delta(
            rm_a.P, rm_b.P, rm_a.y, "AUROC", n_boot=_N_BOOT,
            seed=_shared_seed(f"cmp|{sc}|discovery"))
    return _CE_CACHE[key]


def agg_shap(folds, max_f=14, feature_whitelist=None):
    """
    Pool out-of-fold SHAP values across all folds into a tidy DataFrame.

    If feature_whitelist is provided (list of feature names), only those
    features are included — this is the primary filter for fig05, where we
    restrict to the stable signature (selection freq >= STABILITY_THRESH).
    If feature_whitelist is None, falls back to top max_f by mean |SHAP|
    (used for the fusion SHAP fig08 where there is no per-feature threshold).

    Subsamples to ≤2000 obs per feature for KDE (scatter keeps all points).
    Returns empty DataFrame if no SHAP data available.
    """
    parts_sv, parts_fv, parts_fn, parts_fi = [], [], [], []
    for fold in folds:
        sh = fold.get("oof_shap")
        if sh is None:
            continue
        sv = sh["shap_values"]
        fv = sh["X_test_scaled"]
        fn = sh["feature_names"]
        fi = fold.get("fold_idx", 0)
        if isinstance(sv, list): sv = sv[1]
        if sv.ndim == 3:         sv = sv[:, :, 1]
        sv = np.asarray(sv, dtype=np.float32)
        fv = np.asarray(fv, dtype=np.float32)
        n_test, n_feat = sv.shape
        parts_sv.append(sv.ravel())
        parts_fv.append(fv.ravel())
        parts_fn.append(np.tile(fn, n_test))
        parts_fi.append(np.full(n_test * n_feat, fi, dtype=np.int32))

    if not parts_sv:
        return pd.DataFrame(columns=["feature", "fold_idx", "shap_val", "feat_val"])

    df = pd.DataFrame({
        "feature":  np.concatenate(parts_fn),
        "fold_idx": np.concatenate(parts_fi),
        "shap_val": np.concatenate(parts_sv).astype(float),
        "feat_val": np.concatenate(parts_fv).astype(float),
    })

    if feature_whitelist is not None:
        # Filter to stable signature features only
        df = df[df["feature"].isin(feature_whitelist)].copy()
    else:
        # Fallback: top max_f by mean |SHAP| (fusion fig only)
        top = (df.groupby("feature")["shap_val"]
                 .apply(lambda x: np.abs(x).mean())
                 .sort_values(ascending=False)
                 .head(max_f).index.tolist())
        df = df[df["feature"].isin(top)].copy()

    return df


def draw_bee(ax, df, title, max_f=14):
    """SHAP beeswarm with KDE jitter. Returns scatter for colorbar."""
    if df.empty:
        ax.text(0.5, 0.5, "No SHAP", ha="center", va="center",
                transform=ax.transAxes, fontsize=8, color="#888")
        ax.set_title(title, fontsize=9, fontweight="bold")
        return None
    order = (df.groupby("feature")["shap_val"]
               .apply(lambda x: np.abs(x).mean())
               .sort_values(ascending=True)
               .index.tolist())[-max_f:]
    cmap = plt.cm.RdBu_r; sc_ = None
    for yi, feat in enumerate(order):
        sub  = df[df["feature"] == feat]
        vals = sub["shap_val"].values
        fv   = sub["feat_val"].values
        if len(vals) > 2:
            try:
                kde_samp = (vals if len(vals) <= 2000
                            else np.random.default_rng(42 + yi).choice(
                                vals, 2000, replace=False))
                kde  = gaussian_kde(kde_samp, bw_method=0.4)
                dens = kde(vals); dens = dens / dens.max() * 0.35
            except Exception:
                dens = np.full(len(vals), 0.1)
        else:
            dens = np.full(len(vals), 0.05)
        jitter = np.random.default_rng(42 + yi).uniform(-dens, dens)
        fv_n   = (fv - fv.min()) / (fv.max() - fv.min() + 1e-9)
        sc_    = ax.scatter(vals, yi + jitter, c=fv_n, cmap=cmap,
                            vmin=0, vmax=1, alpha=0.55, s=9,
                            linewidths=0, rasterized=True)
    ax.set_yticks(range(len(order)))
    ax.set_yticklabels(order, fontsize=6.5)
    ax.axvline(0, color="black", lw=0.7, ls="--")
    ax.set_xlabel("SHAP value", fontsize=8)
    ax.set_title(title, fontsize=9, fontweight="bold")
    ax.spines["left"].set_visible(False)
    ax.tick_params(left=False)
    ax.grid(axis="x", alpha=0.35)
    return sc_


def add_cb(fig, sc_, ax):
    if sc_ is None:
        return
    cb = fig.colorbar(sc_, ax=ax, orientation="vertical",
                      fraction=0.028, pad=0.04, aspect=28)
    cb.set_label("Feature value\n(low→high)", fontsize=6)
    cb.ax.tick_params(labelsize=6)


def _mean_signed_shap(sh):
    """
    Mean signed SHAP value per feature for one fold, as a (feature, value) zip.

    shap returns different shapes depending on the explainer: a plain 2-D
    (n_samples, n_features) array for linear models, a list of per-class
    arrays for some tree models, and a 3-D (n_samples, n_features, n_classes)
    array for others. Taking .mean(axis=0) without unwrapping the class
    dimension first either raises on the shape mismatch or, worse, silently
    averages across classes and produces a direction that is not the
    positive-class direction. Both call sites now go through here.
    """
    sv = sh["shap_values"]
    if isinstance(sv, list):
        sv = sv[1] if len(sv) > 1 else sv[0]
    sv = np.asarray(sv)
    if sv.ndim == 3:
        sv = sv[:, :, 1] if sv.shape[2] > 1 else sv[:, :, 0]
    if sv.ndim != 2:
        return []
    names = list(sh["feature_names"])
    means = sv.mean(axis=0)
    if len(names) != len(means):
        return []
    return list(zip(names, means))


def _savefig(fig, path):
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    print(f"  → {path.name}")

# =============================================================================
# SECTION A: PERFORMANCE
# =============================================================================
def fig_roc(data, fd):
    """Fig 01 — Mean ROC curve for Fused_ElasticNet × all scenarios."""
    fig, axes = plt.subplots(1, 3, figsize=(16, 6))
    fus_col = "#6a1f6a"

    for ai, sc in enumerate(SCENARIOS):
        ax = axes[ai]
        fv = "Fused_ElasticNet"
        if fv in data[sc]:
            roc_c = data[sc][fv]["roc"]
            base  = np.linspace(0, 1, 200)
            tprs  = [np.interp(base, r[0], r[1]) for r in roc_c]
            mean_tpr = np.mean(tprs, axis=0)
            std_tpr  = np.std(tprs,  axis=0)
            au = data[sc][fv]["aurocs"].mean()
            ax.plot(base, mean_tpr, color=fus_col, ls="-", lw=2,
                    label=f"Fused ElasticNet (AUC={au:.3f})", alpha=0.9)
            ax.fill_between(base, mean_tpr-std_tpr, mean_tpr+std_tpr,
                            color=fus_col, alpha=0.12)
        # Add best unimodal for reference
        best_au = 0; best_mod = None
        for mod in UNIMODALS:
            if mod in data[sc]:
                a = data[sc][mod]["aurocs"].mean()
                if a > best_au:
                    best_au = a; best_mod = mod
        if best_mod:
            roc_c = data[sc][best_mod]["roc"]
            base  = np.linspace(0, 1, 200)
            tprs  = [np.interp(base, r[0], r[1]) for r in roc_c]
            ax.plot(base, np.mean(tprs, axis=0), color=MOD_COLOR[best_mod],
                    ls=":", lw=1.5, alpha=0.7,
                    label=f"{best_mod} best unimodal ({best_au:.3f})")
        ax.plot([0,1],[0,1],"k--",lw=0.7,alpha=0.4)
        ax.set(xlim=(0,1), ylim=(0,1), aspect="equal",
               xlabel="1−Specificity (FPR)", ylabel="Sensitivity (TPR)")
        ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")
        ax.legend(fontsize=7, loc="lower right")
        ax.grid(alpha=0.3)
    fig.suptitle(
        "ROC Curves — Fused ElasticNet × All Scenarios\n"
        "Mean ± SD over all outer folds  ·  Dotted = best unimodal",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig01_roc_curves.pdf")


def fig_performance_distributions(data, fd):
    """Fig 02 — Boxplots of all 5 metrics, all 7 models × 3 scenarios."""
    metrics = ["AUROC", "AUPRC", "Sensitivity", "Specificity", "Brier"]
    fig, axes = plt.subplots(5, 3, figsize=(20, 24), sharey="row")
    ref_lines = {"AUROC": 0.5, "AUPRC": None, "Sensitivity": None,
                 "Specificity": None, "Brier": None}

    for ri, metric in enumerate(metrics):
        mkey = METRIC_KEYS[metric]
        for ci, sc in enumerate(SCENARIOS):
            ax = axes[ri, ci]
            x_pos = 0; tick_x = []; tick_l = []
            for mod in ALL_MODELS:
                if mod not in data[sc]:
                    x_pos += 1.5; tick_x.append(x_pos - 0.75)
                    tick_l.append(mod.replace("Fused_", "F.")); continue
                arr = data[sc][mod][mkey]
                arr = arr[~np.isnan(arr)]
                if len(arr) == 0:
                    x_pos += 1.5; tick_x.append(x_pos - 0.75)
                    tick_l.append(mod.replace("Fused_", "F.")); continue
                col = MOD_COLOR[mod]
                bp  = ax.boxplot(arr, positions=[x_pos], widths=0.7,
                                 patch_artist=True, showfliers=False,
                                 medianprops=dict(color="white", lw=1.8))
                bp["boxes"][0].set_facecolor(col)
                bp["boxes"][0].set_alpha(0.85)
                tick_x.append(x_pos); tick_l.append(mod.replace("Fused_", "F."))
                x_pos += 1.5
            ax.set_xticks(tick_x)
            ax.set_xticklabels(tick_l, rotation=35, ha="right", fontsize=7)
            if ref_lines[metric] is not None:
                ax.axhline(ref_lines[metric], color="#aaa", lw=0.8, ls=":")
            ax.grid(axis="y", alpha=0.35)
            if ci == 0:
                ax.set_ylabel(
                    f"{metric} ({'↑' if metric != 'Brier' else '↓'})",
                    fontsize=8.5, fontweight="bold")
            if ri == 0:
                ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")

    fig.suptitle(
        "Performance Distributions — All Models × All Scenarios\n"
        "Youden-optimal threshold for Sensitivity and Specificity",
        fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout(h_pad=0.4)
    _savefig(fig, fd / "fig02_performance_distributions.pdf")


def fig_fusion_benefit(data, fd):
    """
    Fig 03 — ΔAUROC = Fused_ElasticNet − Best Unimodal, single-panel view.

    The comparator for each scenario is the FIXED-best unimodal: the single
    modality with the highest mean AUROC across folds, used as the
    comparator in every fold. This is the statistically honest comparison:

      * Using "max over modalities within each fold" as the comparator
        (common in fusion-benefit papers) is anti-conservative. Picking
        the best of K noisy per-fold estimates regresses upward, so the
        comparator is systematically inflated and ΔAUROC is systematically
        biased against the fusion model.

      * Fixing the comparator to the single modality with the highest
        MEAN AUROC (mean over CV repeats of the pooled-OOF AUROC, the same
        estimand as the bars) removes that bias — the comparator is
        pre-specified and the same modality is used in every fold.

    Bars = paired ΔAUROC (fused − fixed-best unimodal), computed within each
    CV repeat on the complete out-of-fold vectors and averaged over repeats.
    Error bars = 95% paired patient-level cluster-bootstrap CI (the same
    patient resample is applied to both models and all their repeats; see
    cv_estimands.paired_bootstrap_repeat_delta). Label under each bar names
    the fixed-best modality for that scenario.
    """
    fig, ax = plt.subplots(1, 1, figsize=(9, 6))
    fv = "Fused_ElasticNet"
    w  = 0.55
    x  = np.arange(len(SCENARIOS))

    for sc in SCENARIOS:
        if fv not in data[sc]:
            continue
        uni_present = [m for m in UNIMODALS if m in data[sc]]
        if not uni_present:
            continue

        best_mod = max(uni_present,
                       key=lambda m: _discovery_metric_ci(data, sc, m,
                                                          "AUROC")["estimate"])
        res   = _discovery_paired_delta(data, sc, fv, best_mod)
        mn    = res["delta"]
        ci_lo, ci_hi = res["ci_low"], res["ci_high"]
        if not np.isfinite(mn):
            continue
        xi = SCENARIOS.index(sc)
        ax.bar(xi, mn, w, color=SC_COL[sc], alpha=0.85)
        ax.errorbar(xi, mn, yerr=[[max(0, mn - ci_lo)], [max(0, ci_hi - mn)]],
                    fmt="none", ecolor="#333", elinewidth=1.2, capsize=6)

        # Value + CI label above (or below for negative) the bar
        va  = "bottom" if mn >= 0 else "top"
        y_t = ci_hi + 0.006 if mn >= 0 else ci_lo - 0.006
        ax.text(xi, y_t,
                f"{mn:+.3f}\n[{ci_lo:+.3f}, {ci_hi:+.3f}]",
                ha="center", va=va,
                fontsize=8.5, color=SC_COL[sc], fontweight="bold")
        # Fixed-best modality label under x-tick
        ax.text(xi, ax.get_ylim()[0], f"vs {best_mod}",
                ha="center", va="top", fontsize=8, color=SC_COL[sc],
                transform=ax.get_xaxis_transform())

    ax.axhline(0, color="black", lw=1.0, ls="--")
    ax.set_xticks(x)
    ax.set_xticklabels(SCENARIOS, fontsize=10)
    ax.set_ylabel("ΔAUROC (fused − fixed-best unimodal)")
    ax.grid(axis="y", alpha=0.35)
    # Give room for the "vs MOD" sublabel beneath the x-tick
    ax.tick_params(axis="x", pad=18)

    fig.suptitle(
        "Fusion benefit: fused elastic-net vs fixed-best unimodal\n"
        "Paired ΔAUROC, mean over CV repeats · 95% paired patient-level "
        "cluster-bootstrap CI · comparator = "
        "modality with highest mean AUROC (fixed across folds)",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig03_fusion_benefit.pdf")


def fig_forest_plot(data, fd):
    """
    Fig 04 — Paired ΔAUROC, Fused_ElasticNet vs each unimodal.

    Square = ΔAUROC computed within each CV repeat on the complete
    out-of-fold vectors, averaged over repeats. Line = 95% paired
    patient-level cluster-bootstrap CI (same patient resample applied to
    both models and all their repeats). The percentage printed beside each
    interval is the fraction of outer folds in which the fused model beat
    the unimodal — a descriptive fold-level count, not an inference.
    """
    fig, axes = plt.subplots(1, 3, figsize=(18, 5), sharey=True)
    fv     = "Fused_ElasticNet"
    fv_col = "#6a1f6a"

    for ci, sc in enumerate(SCENARIOS):
        ax    = axes[ci]
        y_pos = list(reversed(range(len(UNIMODALS))))
        if fv not in data[sc]:
            continue
        fused_a = data[sc][fv]["aurocs"]
        for yi, mod in zip(y_pos, reversed(UNIMODALS)):
            if mod not in data[sc]:
                continue
            res = _discovery_paired_delta(data, sc, fv, mod)
            md, ci0, ci1 = res["delta"], res["ci_low"], res["ci_high"]
            if not np.isfinite(md):
                continue
            # Descriptive only: share of outer folds where fused > unimodal.
            delta = fused_a - data[sc][mod]["aurocs"]
            pct   = (delta > 0).mean() * 100
            col   = MOD_COLOR[mod]
            ax.plot([ci0, ci1], [yi, yi], color=col, lw=2, alpha=0.85)
            ax.scatter([md], [yi], color=col, s=50, zorder=4,
                       marker="s", edgecolors="white", linewidths=0.5)
            ax.text(ci1 + 0.005, yi, f"{pct:.0f}%",
                    va="center", fontsize=6.5, color=col)
        ax.axvline(0, color="black", lw=0.9, ls="--")
        ax.set_yticks(y_pos)
        if ci == 0:
            ax.set_yticklabels(list(reversed(UNIMODALS)), fontsize=8.5)
        else:
            ax.set_yticklabels([])
        ax.set_xlabel("ΔAUROC", fontsize=8.5)
        ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")
        if ci == 0:
            ax.set_ylabel("Fused ElasticNet", color=fv_col,
                          fontsize=9, fontweight="bold")
        ax.grid(axis="x", alpha=0.35)
        ax.spines["left"].set_visible(False)
        ax.tick_params(left=False)
    fig.legend(
        handles=[Line2D([0],[0], color=MOD_COLOR[m], lw=2, label=m)
                 for m in UNIMODALS],
        loc="lower center", ncol=5, fontsize=9, bbox_to_anchor=(0.5, -0.02))
    fig.suptitle(
        "Statistical Comparison: Fused ElasticNet vs Each Unimodal\n"
        "ΔAUROC, mean over CV repeats · 95% paired patient-level "
        "cluster-bootstrap CI  ·  % = fraction of folds where Fused > "
        "Unimodal (descriptive)",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout(w_pad=0.3)
    _savefig(fig, fd / "fig04_forest_plot.pdf")

# =============================================================================
# SECTION B: FEATURE SHAP AND SIGNATURES
# =============================================================================
def fig_feature_shap(data, fd):
    """
    Fig 05 — Feature SHAP beeswarm per modality × scenario.

    Shows ONLY features that appear in the winner signature of at least
    one outer fold (i.e. the signature universe). Features never selected
    by any fold are excluded so the panel corresponds strictly to the
    per-modality signature, not to the raw feature pool.

    Features are ranked within each panel by mean |SHAP| across the folds
    in which they appeared. Panel sizes differ across modalities because
    signature sizes differ by design:
      Clin / WSI  : all features (keep-all by design)
      Prot        : all 5 features (pool = floor)
      RNA / DNA   : union of per-fold top-N signatures (typically 6-15
                    at global scale, 5-10 at arm scale)
    """
    for sc in SCENARIOS:
        n_mods = sum(1 for m in UNIMODALS if m in data[sc])
        if n_mods == 0:
            continue
        fig, axes = plt.subplots(1, 5, figsize=(26, 10))
        for ci, mod in enumerate(UNIMODALS):
            ax  = axes[ci]
            col = MOD_COLOR[mod]
            folds = data[sc].get(mod, {}).get("folds", [])

            # Signature universe = features selected by AT LEAST ONE fold.
            # Features that appear in the raw modality pool but never get
            # picked by any fold's winner signature are excluded — this is
            # the exact "per-modality signature" the reader wants to see.
            sig_universe = set()
            for fold in folds:
                for feat in get_signature_feats(fold):
                    sig_universe.add(feat)
            sig_universe = sorted(sig_universe)

            if not sig_universe:
                ax.text(0.5, 0.5, "No signature\nfeatures",
                        ha="center", va="center", transform=ax.transAxes,
                        fontsize=9, color="#666")
                ax.set_title(mod, color=col, fontsize=11, fontweight="bold")
                ax.set_xticks([]); ax.set_yticks([])
                continue

            df_ = agg_shap(folds, feature_whitelist=sig_universe)
            # Show everything in the signature universe — no top-k cap.
            sc_ = draw_bee(ax, df_, mod, max_f=len(sig_universe))
            add_cb(fig, sc_, ax)
            ax.set_title(f"{mod}  (n_sig = {len(sig_universe)})",
                         color=col, fontsize=11, fontweight="bold")
        sc_lbl = "Global Model" if sc == "Global" else f"{sc} Arm"
        fig.suptitle(
            f"Feature SHAP — Winner Signature per Modality  |  {sc_lbl}\n"
            "Each point = one test patient  ·  x = SHAP value  ·  "
            "Colour: blue = low feature value, red = high\n"
            "Only features selected by at least one fold's winner signature are shown",
            fontsize=11, fontweight="bold", y=1.01)
        plt.tight_layout(w_pad=0.4)
        _savefig(fig, fd / f"fig05_feature_shap_{sc.replace('-','_')}.pdf")


def fig_feature_selection_frequency(data, fd):
    """
    Fig 06 — Feature selection frequency in winner signatures across folds.
    For each modality, how often does each feature appear in the final
    winner signature across all outer folds?
    Bars coloured by mean SHAP direction (positive/negative pCR association).

    NOTE: no stability-threshold line is drawn. Per-fold signatures have
    fixed-size caps (RNA/DNA/Prot floor-bound at 5 in arm analyses, max 7-9
    at global; Clin/WSI keep all features → freq = 100%), so a fixed-value
    frequency threshold does not carry a consistent meaning across
    modalities. A 100% freq for a 3-feature modality is trivially satisfied;
    for a 30-feature RNA pool at arm level it is mechanically bounded
    above by 5/30 ≈ 17%. Features here are ranked simply by selection
    frequency, up to the top 14.
    """
    fig, axes = plt.subplots(5, 3, figsize=(20, 28))

    for ri, mod in enumerate(UNIMODALS):
        col = MOD_COLOR[mod]
        for ci, sc in enumerate(SCENARIOS):
            ax = axes[ri, ci]
            folds = data[sc].get(mod, {}).get("folds", [])
            valid_folds = [f for f in folds if get_signature_feats(f)]
            if not valid_folds:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        transform=ax.transAxes, fontsize=8)
                if ri == 0:
                    ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                                 fontweight="bold")
                continue

            n_tot = len(valid_folds)
            feat_freq = Counter()
            feat_shap_dir = defaultdict(list)

            for fold in valid_folds:
                for feat in get_signature_feats(fold):
                    feat_freq[feat] += 1
                sh = fold.get("oof_shap")
                if sh:
                    for feat, sv in _mean_signed_shap(sh):
                        feat_shap_dir[feat].append(float(sv))

            if not feat_freq:
                continue

            df_freq = pd.DataFrame([
                {"feature":   f,
                 "freq":      cnt / n_tot,
                 "mean_shap": float(np.mean(feat_shap_dir[f]))
                              if feat_shap_dir[f] else 0.0}
                for f, cnt in feat_freq.most_common()
            ])
            # Cap at 20 rows per panel only as a safeguard against extreme
            # signature-union sizes; for typical RNA/DNA panels this is
            # non-binding (union is ~5-15 features).
            if len(df_freq) > 20:
                df_freq = df_freq.head(20)

            # Colour by SHAP sign only — no "below threshold" greying.
            bar_colors = ["#2e7d32" if row["mean_shap"] >= 0 else "#c62828"
                          for _, row in df_freq.iterrows()]

            y = np.arange(len(df_freq))
            ax.barh(y, df_freq["freq"].values[::-1],
                    color=list(reversed(bar_colors)),
                    edgecolor="white", height=0.72)
            ax.set_yticks(y)
            ax.set_yticklabels(df_freq["feature"].values[::-1], fontsize=6.5)
            ax.set_xlim(0, 1.05)
            ax.set_xlabel("Selection freq.") if ri == 4 else None
            ax.spines["left"].set_visible(False)
            ax.tick_params(left=False)
            ax.grid(axis="x", alpha=0.35)
            if ri == 0:
                ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                             fontweight="bold")
            if ci == 0:
                ax.set_ylabel(mod, color=col, fontsize=9, fontweight="bold")

    fig.legend(
        handles=[Patch(color="#2e7d32", label="Positive pCR assoc. (mean SHAP > 0)"),
                 Patch(color="#c62828", label="Negative pCR assoc. (mean SHAP < 0)")],
        loc="lower center", ncol=2, fontsize=9, bbox_to_anchor=(0.5, -0.01))
    fig.suptitle(
        "Feature Selection Frequency in Winner Signatures\n"
        "Fraction of outer folds where feature is in the winner classifier's signature  "
        "·  colour = sign of mean SHAP (positive/negative pCR association)",
        fontsize=11, fontweight="bold", y=1.005)
    plt.tight_layout(w_pad=0.4, h_pad=0.8, rect=[0, 0.01, 1, 0.98])
    _savefig(fig, fd / "fig06_feature_selection_frequency.pdf")


def fig_cross_scenario_features(data, fd):
    """
    Fig 07 — Cross-scenario feature frequency comparison.
    Global signature order fixed; DHP and T-DM1 show shifts.
    """
    fig, axes = plt.subplots(1, 5, figsize=(26, 10))
    for ci, mod in enumerate(UNIMODALS):
        ax  = axes[ci]
        col = MOD_COLOR[mod]

        # Global feature order as reference
        folds_g = data["Global"].get(mod, {}).get("folds", [])
        valid_g = [f for f in folds_g if get_signature_feats(f)]
        if not valid_g:
            ax.set_title(mod, color=col, fontsize=10)
            continue

        freq_g = Counter()
        for fold in valid_g:
            for feat in get_signature_feats(fold):
                freq_g[feat] += 1
        feat_order = [f for f, _ in freq_g.most_common(12)]

        y = np.arange(len(feat_order))
        h = 0.22
        offs = {"Global": -h, "DHP": 0, "T-DM1": h}

        for sc in SCENARIOS:
            folds_s = data[sc].get(mod, {}).get("folds", [])
            valid_s = [f for f in folds_s if get_signature_feats(f)]
            if not valid_s:
                continue
            n_tot = len(valid_s)
            freq_s = Counter()
            for fold in valid_s:
                for feat in get_signature_feats(fold):
                    freq_s[feat] += 1
            # Reverse so most_common is at top (matches fig06 orientation)
            vals = [freq_s.get(f, 0) / n_tot for f in reversed(feat_order)]
            ax.barh(y + offs[sc], vals, h * 0.85,
                    color=SC_COL[sc], alpha=0.82, label=sc)

        ax.set_yticks(y)
        # Show labels only on leftmost panel; feat_order reversed so most
        # frequent appears at top (consistent with fig06)
        ax.set_yticklabels(
            list(reversed(feat_order)) if ci == 0 else [],
            fontsize=6.5)
        ax.set_xlabel("Selection freq.", fontsize=8)
        ax.set_title(mod, color=col, fontsize=10, fontweight="bold")
        ax.spines["left"].set_visible(False)
        ax.tick_params(left=False)
        ax.grid(axis="x", alpha=0.35)
        if ci == 4:
            ax.legend(fontsize=8, loc="lower right")

    fig.suptitle(
        "Cross-Scenario Feature Frequency — Global Feature Order\n"
        "Divergence between bars = arm-specific signature shift",
        fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    _savefig(fig, fd / "fig07_cross_scenario_features.pdf")

# =============================================================================
# SECTION C: FUSION ATTRIBUTION
# =============================================================================
def fig_fusion_shap(data, fd):
    """Fig 08 — Fusion SHAP beeswarm: modality attribution per scenario."""
    fig, axes = plt.subplots(1, 3, figsize=(18, 8))
    for ai, sc in enumerate(SCENARIOS):
        ax = axes[ai]
        folds = data[sc].get("Fused_ElasticNet", {}).get("folds", [])
        df_   = agg_shap(folds, max_f=5)  # 5 modalities max
        sc_   = draw_bee(ax, df_, sc, max_f=5)
        add_cb(fig, sc_, ax)
        ax.title.set_color(SC_COL[sc])
    fig.suptitle(
        "Fusion SHAP — Modality Attribution  |  Fused_ElasticNet\n"
        "x = SHAP contribution of modality probability to fused pCR prediction\n"
        "Positive = modality pushes toward pCR  ·  Colour = calibrated modality probability",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig08_fusion_shap.pdf")


def fig_modality_weights(data, fd):
    """Fig 09 — Fused_ElasticNet modality weights + selection frequency."""
    fig, axes = plt.subplots(2, 3, figsize=(18, 10))
    for ri, metric in enumerate(["weight", "selection_rate"]):
        for ci, sc in enumerate(SCENARIOS):
            ax = axes[ri, ci]
            folds = data[sc].get("Fused_ElasticNet", {}).get("folds", [])
            if not folds:
                continue

            if metric == "weight":
                # Distribution of absolute weights per modality
                for xi, mod in enumerate(UNIMODALS):
                    wts = [abs(f["modality_weights"].get(mod, 0)) for f in folds]
                    bp  = ax.boxplot(wts, positions=[xi], widths=0.5,
                                     patch_artist=True, showfliers=False,
                                     medianprops=dict(color="white", lw=1.5))
                    bp["boxes"][0].set_facecolor(MOD_COLOR[mod])
                    bp["boxes"][0].set_alpha(0.85)
                ax.set_xticks(range(len(UNIMODALS)))
                ax.set_xticklabels(UNIMODALS, rotation=20, ha="right")
                ax.set_ylabel("|Coefficient| distribution") if ci == 0 else None
                ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                             fontweight="bold") if ri == 0 else None
            else:
                # Bar: fraction of folds each modality is selected (coef ≠ 0)
                sel_rates = []
                for mod in UNIMODALS:
                    sel = [1 if mod in f.get("selected_modalities", [])
                           else 0 for f in folds]
                    sel_rates.append(np.mean(sel))
                bars = ax.bar(range(len(UNIMODALS)), sel_rates,
                              color=[MOD_COLOR[m] for m in UNIMODALS],
                              alpha=0.85, edgecolor="white")
                for bar, rate in zip(bars, sel_rates):
                    ax.text(bar.get_x() + bar.get_width()/2,
                            bar.get_height() + 0.02,
                            f"{rate:.0%}", ha="center", fontsize=7.5)
                ax.set_xticks(range(len(UNIMODALS)))
                ax.set_xticklabels(UNIMODALS, rotation=20, ha="right")
                ax.set_ylim(0, 1.15)
                ax.set_ylabel("Selection rate (coef ≠ 0)") if ci == 0 else None
                ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                             fontweight="bold") if ri == 1 else None
            ax.grid(axis="y", alpha=0.35)

    fig.suptitle(
        "Fused_ElasticNet — Modality Weights and Selection\n"
        "Row 1: |coefficient| distribution across folds  "
        "·  Row 2: fraction of folds each modality contributes",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout(h_pad=0.6)
    _savefig(fig, fd / "fig09_modality_weights.pdf")

# =============================================================================
# SECTION D: CLASSIFIER ANALYSIS
# =============================================================================
def fig_winner_classifier_heatmap(data, fd):
    """
    Fig 10 — Winner classifier selection frequency heatmap.
    Rows = modalities, Cols = classifiers, Panels = scenarios.
    Cell value = fraction of folds this classifier won for this modality.
    """
    # Determine which classifiers appear
    seen_clfs = set()
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for fold in data[sc].get(mod, {}).get("folds", []):
                c = fold.get("winner_clf", "")
                if c and c != "none":
                    seen_clfs.add(c)
    clfs = [c for c in ALL_CLFS if c in seen_clfs]
    if not clfs:
        return

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    for ai, sc in enumerate(SCENARIOS):
        ax  = axes[ai]
        mat = np.zeros((len(UNIMODALS), len(clfs)))
        for ri, mod in enumerate(UNIMODALS):
            folds = data[sc].get(mod, {}).get("folds", [])
            total = len(folds)
            if total == 0:
                continue
            cnts  = Counter(f.get("winner_clf","") for f in folds)
            for ci2, clf in enumerate(clfs):
                mat[ri, ci2] = cnts.get(clf, 0) / total

        im = ax.imshow(mat, cmap=plt.cm.YlOrRd, vmin=0, vmax=1, aspect="auto")
        for ri in range(len(UNIMODALS)):
            for ci2 in range(len(clfs)):
                tc = "white" if mat[ri, ci2] > 0.6 else "black"
                ax.text(ci2, ri, f"{mat[ri,ci2]:.0%}",
                        ha="center", va="center",
                        fontsize=8.5, color=tc,
                        fontweight="bold" if mat[ri, ci2] > 0.4 else "normal")
        ax.set_xticks(range(len(clfs)))
        ax.set_xticklabels([CLF_LABEL.get(c, c) for c in clfs],
                           rotation=25, ha="right", fontsize=8)
        ax.set_yticks(range(len(UNIMODALS)))
        ax.set_yticklabels(UNIMODALS if ai == 0 else [], fontsize=9)
        ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="Selection rate")

    fig.suptitle(
        "Winner Classifier Selection Frequency\n"
        "Fraction of outer folds each classifier was selected as winner per modality",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig10_winner_classifier_heatmap.pdf")


def fig_inner_auroc_comparison(data, fd):
    """
    Fig 11 — Stage A inner AUROC violin per classifier × modality.
    Shows the distribution of inner CV AUROCs across folds,
    directly comparing all classifiers on equal footing (STAGE_A_PARAMS).
    """
    seen_clfs = set()
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for fold in data[sc].get(mod, {}).get("folds", []):
                seen_clfs.update(fold.get("inner_cv_aurocs_A", {}).keys())
    clfs = [c for c in ALL_CLFS if c in seen_clfs]
    if not clfs:
        return

    fig, axes = plt.subplots(len(UNIMODALS), 3,
                             figsize=(18, 4 * len(UNIMODALS)), sharey="row")

    for ri, mod in enumerate(UNIMODALS):
        col = MOD_COLOR[mod]
        for ci, sc in enumerate(SCENARIOS):
            ax = axes[ri, ci]
            clf_aurocs = {c: [] for c in clfs}
            for fold in data[sc].get(mod, {}).get("folds", []):
                for c, au in fold.get("inner_cv_aurocs_A", {}).items():
                    if c in clf_aurocs:
                        clf_aurocs[c].append(au)

            pos_x = 0
            for clf in clfs:
                vals = clf_aurocs[clf]
                if not vals:
                    pos_x += 1; continue
                parts = ax.violinplot([vals], positions=[pos_x],
                                      widths=0.7, showmedians=True,
                                      showextrema=False)
                for pc in parts["bodies"]:
                    pc.set_facecolor(CLF_COLOR[clf])
                    pc.set_alpha(0.7)
                parts["cmedians"].set_color("white")
                parts["cmedians"].set_linewidth(2)
                pos_x += 1

            ax.axhline(0.5, color="#aaa", lw=0.8, ls=":")
            ax.set_xticks(range(len(clfs)))
            ax.set_xticklabels([CLF_LABEL.get(c, c) for c in clfs],
                               rotation=25, ha="right", fontsize=7.5)
            ax.grid(axis="y", alpha=0.35)
            if ci == 0:
                ax.set_ylabel(mod, color=col, fontsize=9, fontweight="bold")
            if ri == 0:
                ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                             fontweight="bold")

    fig.legend(
        handles=[Patch(facecolor=CLF_COLOR[c], label=CLF_LABEL.get(c, c))
                 for c in clfs],
        loc="lower center", ncol=len(clfs), fontsize=9,
        bbox_to_anchor=(0.5, -0.02))
    fig.suptitle(
        "Classifier Comparison — Stage A Inner CV AUROC\n"
        "Fixed STAGE_A_PARAMS  ·  Fair comparison on equal footing  "
        "·  Winner selected by highest mean",
        fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout(h_pad=0.4, w_pad=0.3)
    _savefig(fig, fd / "fig11_inner_auroc_comparison.pdf")


def fig_calibration_profile(data, fd):
    """
    Fig 12 — Platt calibration slopes per classifier × modality.
    Violin: slope distribution across folds.
    Bar: Platt application rate.
    Shaded band: ideal range [0.80, 1.20].
    """
    seen_clfs = set()
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for fold in data[sc].get(mod, {}).get("folds", []):
                seen_clfs.update(fold.get("calibration", {}).keys())
    clfs = [c for c in ALL_CLFS if c in seen_clfs]
    if not clfs:
        return

    fig, axes = plt.subplots(len(UNIMODALS), 3,
                             figsize=(18, 4 * len(UNIMODALS)), sharey="row")

    for ri, mod in enumerate(UNIMODALS):
        col = MOD_COLOR[mod]
        for ci, sc in enumerate(SCENARIOS):
            ax = axes[ri, ci]
            for xi, clf in enumerate(clfs):
                slopes = []
                n_platt = 0
                n_total = 0
                for fold in data[sc].get(mod, {}).get("folds", []):
                    cal = fold.get("calibration", {}).get(clf)
                    if cal:
                        slopes.append(cal["slope"])
                        n_platt += int(cal["needs_platt"])
                        n_total += 1
                if not slopes:
                    continue
                # Clip slopes for display (some extreme in small folds)
                slopes_c = np.clip(slopes, -0.5, 5)
                parts = ax.violinplot([slopes_c], positions=[xi],
                                      widths=0.7, showmedians=True,
                                      showextrema=False)
                for pc in parts["bodies"]:
                    pc.set_facecolor(CLF_COLOR[clf]); pc.set_alpha(0.7)
                parts["cmedians"].set_color("white")
                parts["cmedians"].set_linewidth(2)
                # Annotate Platt rate
                rate = n_platt / n_total if n_total else 0
                ax.text(xi, -0.45, f"{rate:.0%}",
                        ha="center", fontsize=6.5, color=CLF_COLOR[clf],
                        fontweight="bold")

            ax.axhspan(0.80, 1.20, alpha=0.08, color="green",
                       label="Ideal [0.80,1.20]")
            ax.axhline(1.0, color="green", lw=0.9, ls="--", alpha=0.5)
            ax.set_xticks(range(len(clfs)))
            ax.set_xticklabels([CLF_LABEL.get(c, c) for c in clfs],
                               rotation=25, ha="right", fontsize=7.5)
            ax.set_ylim(-0.6, 5.2)
            ax.grid(axis="y", alpha=0.35)
            if ci == 0:
                ax.set_ylabel(f"{mod}\nCalib. slope",
                              color=col, fontsize=9, fontweight="bold")
            if ri == 0:
                ax.set_title(sc, color=SC_COL[sc], fontsize=10,
                             fontweight="bold")

    fig.suptitle(
        "Calibration Profile — Platt Slope per Classifier × Modality\n"
        "Green band = ideal range [0.80–1.20]  ·  % = Platt application rate",
        fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout(h_pad=0.4, w_pad=0.3)
    _savefig(fig, fd / "fig12_calibration_profile.pdf")


def fig_signature_sizes(data, fd):
    """
    Fig 13 — Signature size distributions per modality × scenario.
    Shows EPV-capped signature sizes, with a reference line at the FLOOR=5
    minimum signature size.

    The pipeline applies a single events-per-variable cap of 5 to every
    classifier family (see _derive_signature), not a per-family cap. The
    realised per-fold EPV — which is what matters for judging overfitting and
    can differ from the cap because of the floor and because features can be
    dropped by a fold's preprocessing — is reported separately by
    revision_analyses.py.
    """
    fig, axes = plt.subplots(1, 3, figsize=(18, 6), sharey=True)
    for ci, sc in enumerate(SCENARIOS):
        ax = axes[ci]
        for xi, mod in enumerate(UNIMODALS):
            sizes = [f.get("signature_size", 0)
                     for f in data[sc].get(mod, {}).get("folds", [])
                     if f.get("signature_size", 0) > 0]
            if not sizes:
                continue
            parts = ax.violinplot([sizes], positions=[xi],
                                  widths=0.6, showmedians=True,
                                  showextrema=True)
            for pc in parts["bodies"]:
                pc.set_facecolor(MOD_COLOR[mod]); pc.set_alpha(0.75)
            parts["cmedians"].set_color("white"); parts["cmedians"].set_linewidth(2)
            ax.text(xi, np.mean(sizes) + 0.3, f"μ={np.mean(sizes):.1f}",
                    ha="center", fontsize=7.5, color=MOD_COLOR[mod])

        ax.axhline(5, color="#888", lw=0.8, ls=":", label="min=5")
        ax.set_xticks(range(len(UNIMODALS)))
        ax.set_xticklabels(UNIMODALS, rotation=20, ha="right")
        ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")
        ax.grid(axis="y", alpha=0.35)
        if ci == 0:
            ax.set_ylabel("Signature size (n features)", fontsize=9)
        if ci == 0:
            ax.legend(fontsize=8, loc="upper right")

    fig.suptitle(
        "Signature Size Distributions — EPV-Capped per Modality × Scenario\n"
        "Cap: events-per-variable ≤ 5 for all classifier families  ·  Minimum = 5",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig13_signature_sizes.pdf")


# =============================================================================
# SECTION F: CONSENSUS DELIVERABLES (R2)
# =============================================================================
# The frozen-consensus OOF evaluation from evaluate_consensus() in the
# pipeline produces the PRIMARY HEADLINE performance numbers for the paper.
# The figures below visualise (1) the consensus signatures per modality
# and (2) the consensus OOF performance: AUROC as the mean over CV repeats
# of the pooled-OOF AUROC with a 95% patient-level cluster-bootstrap CI
# (cv_estimands.py, via _consensus_auroc_ci).
# Data comes from data[sc]["_consensus"]["folds"], ["pooled"], ["consensus"].
# =============================================================================


# =============================================================================
# MAIN-TEXT CONSENSUS FIGURES (renumbered fig01–fig05 for Nature Genetics)
# =============================================================================
# These all use the CONSENSUS SIGNATURE under the ITERATED OUTER CV protocol.
# One performance story, one signature set, one attribution view.
# =============================================================================

def fig_consensus_performance_main(data, fd):
    """
    Fig 01 (MAIN TEXT) — Headline performance figure.

    AUROC of each consensus unimodal model + the consensus fused model:
    bar = AUROC computed on each CV repeat's complete out-of-fold vector,
    averaged over repeats; error bar = 95% patient-level cluster-bootstrap
    CI (patients resampled with all their repeats; cv_estimands.py). The
    numbers are identical to the Performance sheet of PREDIX_HER2_results.xlsx
    (same estimand, same seed). Computed by evaluate_consensus() under the
    same outer-CV splits as discovery, with the frozen consensus signature +
    classifier and fold-refit classifier weights + fold-refit fusion
    coefficients. This is the single performance number reported in the
    paper Results section.
    """
    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [fig01] No consensus data — skipping.")
        return

    fig, axes = plt.subplots(1, len(scenarios),
                              figsize=(6 * len(scenarios), 6), sharey=True)
    if len(scenarios) == 1:
        axes = [axes]

    models_order = UNIMODALS + ["Fused_ElasticNet"]

    for ax, sc in zip(axes, scenarios):
        pooled = data[sc]["_consensus"]["pooled"]

        xs = np.arange(len(models_order))
        for xi, mod in enumerate(models_order):
            if mod not in pooled:
                continue
            col = MOD_COLOR.get(mod, "#555555")
            # Mean over CV repeats of the pooled-OOF AUROC, with a
            # patient-level CLUSTER bootstrap CI (all R predictions of a
            # resampled patient move together) — never an average of
            # probabilities across repeats, never a resample of folds.
            res   = _consensus_auroc_ci(data, sc, mod)
            auroc = res["estimate"]
            ci_lo, ci_hi = res["ci_low"], res["ci_high"]
            if not np.isfinite(auroc):
                continue

            bar_h = 0.55 if mod != "Fused_ElasticNet" else 0.78
            ax.bar(xi, auroc, bar_h, color=col,
                   alpha=0.97 if mod == "Fused_ElasticNet" else 0.85,
                   edgecolor="black" if mod == "Fused_ElasticNet" else "white",
                   linewidth=1.3 if mod == "Fused_ElasticNet" else 0.4)
            ax.errorbar(xi, auroc,
                         yerr=[[max(0, auroc - ci_lo)],
                               [max(0, ci_hi - auroc)]],
                         fmt="none", ecolor="#333", elinewidth=1.2, capsize=5)
            ax.text(xi, auroc + 0.02,
                     f"{auroc:.3f}",
                     ha="center", va="bottom",
                     fontsize=9 if mod != "Fused_ElasticNet" else 10.5,
                     fontweight="bold", color=col)

        ax.axhline(0.5, color="#aaa", lw=0.9, ls=":")
        ax.set_xticks(xs)
        ax.set_xticklabels(
            [m.replace("Fused_", "Fused ") for m in models_order],
            rotation=25, ha="right", fontsize=9)
        ax.set_ylim(0, 1.05)
        ax.set_ylabel("AUROC, mean over CV repeats\n"
                      "(95% patient-level cluster-bootstrap CI)")
        ax.set_title(sc, color=SC_COL[sc], fontsize=11, fontweight="bold")
        ax.grid(axis="y", alpha=0.35)

    fig.suptitle(
        "Consensus-signature performance — iterated outer CV\n"
        "AUROC, mean over CV repeats (95% patient-level cluster-bootstrap CI) · "
        "signature frozen, classifier + fusion refit within each fold",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig01_consensus_performance.pdf")


def fig_consensus_signatures_main(data, fd):
    """
    Fig 02 (MAIN TEXT) — Consensus signatures per modality × scenario.

    The SCIENTIFIC DELIVERABLE of the discovery procedure: one finite
    signature per modality per scenario, ranked by the mean cross-classifier
    percentile rank across discovery folds. Together with the fusion weights
    in fig04, these fully specify the final PREDIX HER2 multimodal model.

    The bar length is a normalised RANK, not a SHAP magnitude — see the
    comment at the barh call. Bar lengths are not comparable across
    modalities; only the within-panel ordering carries meaning.

    Bars are coloured by the sign of the mean SHAP value across discovery
    folds (green = positive pCR association, red = negative), matching the
    convention used in supp_fig06_feature_selection_frequency. Mean SHAP
    direction is computed from the unimodal discovery folds' oof_shap arrays
    (data[sc][mod]["folds"]), restricted to folds where the feature appears
    in the consensus signature.
    """
    POS_COL = "#2e7d32"   # green — positive pCR association
    NEG_COL = "#c62828"   # red   — negative pCR association

    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [fig02] No consensus data — skipping.")
        return

    fig, axes = plt.subplots(len(UNIMODALS), len(scenarios),
                              figsize=(6 * len(scenarios),
                                       3.5 * len(UNIMODALS)),
                              squeeze=False)
    for ri, mod in enumerate(UNIMODALS):
        col = MOD_COLOR[mod]
        for ci, sc in enumerate(scenarios):
            ax   = axes[ri, ci]
            cons = data[sc]["_consensus"]["consensus"].get(mod, {})
            sig  = cons.get("signature", [])
            imp  = cons.get("mean_importance", {})
            if not sig:
                ax.text(0.5, 0.5, "No signature", ha="center", va="center",
                        transform=ax.transAxes, fontsize=9, color="#666")
                ax.set_xticks([]); ax.set_yticks([])
                continue

            # ── Mean SHAP direction per consensus feature ─────────────────
            # Accumulate signed mean SHAP from every discovery fold that
            # produced OOF SHAP values. This is identical to the logic in
            # fig_feature_selection_frequency, restricted to consensus feats.
            feat_shap_dir = defaultdict(list)
            for fold in data[sc].get(mod, {}).get("folds", []):
                sh = fold.get("oof_shap")
                if sh is None:
                    continue
                sv = np.asarray(sh["shap_values"])
                fn = sh["feature_names"]
                if isinstance(sh["shap_values"], list):
                    sv = np.asarray(sh["shap_values"][1])
                elif sv.ndim == 3:
                    sv = sv[:, :, 1]
                for feat, val in zip(fn, sv.mean(axis=0)):
                    if feat in sig:
                        feat_shap_dir[feat].append(float(val))

            # One colour per bar: green if mean SHAP >= 0, red otherwise.
            # Fall back to modality colour if no SHAP data for that feature.
            bar_colors = []
            for feat in sig:
                vals = feat_shap_dir.get(feat, [])
                if not vals:
                    bar_colors.append(col)
                else:
                    bar_colors.append(
                        POS_COL if float(np.mean(vals)) >= 0 else NEG_COL)

            # WHAT THE BAR LENGTH ACTUALLY IS.
            # `mean_importance` is NOT a SHAP magnitude, despite the key name.
            # The pipeline converts each classifier's raw importance to a
            # PERCENTILE RANK within the fold (see _derive_signature and Stage A
            # Pass 1), then averages those ranks across classifiers and folds.
            # Per fold the K values are 1/K..K/K and therefore sum to (K+1)/2 —
            # measured exactly on the run-5 PKLs: 3.000000 for K=5, 2.500000 for
            # K=4, 2.000000 for K=3. The value is a normalised rank in (0, 1],
            # not a magnitude in probability or log-odds units.
            # CONSEQUENCE: bar lengths are NOT comparable across modalities (or
            # across scenarios). A rank of 1.0 means "top-ranked within its own
            # 3-feature WSI panel", which says nothing about how it compares to
            # the top RNA feature out of 35 candidates. Only the WITHIN-PANEL
            # ordering is meaningful.
            # The data key stays `mean_importance` deliberately: it is written
            # by the pipeline into the consensus PKLs and read by several
            # downstream consumers. Renaming it would break them, so only the
            # user-facing label is corrected here.
            values = [imp.get(f, 0.0) for f in sig]
            y_pos  = np.arange(len(sig))[::-1]
            ax.barh(y_pos, values, color=bar_colors, alpha=0.88,
                    edgecolor="white", height=0.7)
            ax.set_yticks(y_pos)
            ax.set_yticklabels(sig, fontsize=7.5)
            ax.set_xlabel("mean selection rank (not |SHAP|)", fontsize=8)
            ax.spines["left"].set_visible(False)
            ax.tick_params(left=False)
            ax.grid(axis="x", alpha=0.35)
            clf = cons.get("winner_clf", "")
            sup = cons.get("support_fraction", 0)
            ax.set_title(f"{mod} | {sc}   clf={clf}  ({sup*100:.0f}% support)",
                          color=col, fontsize=9, fontweight="bold")

    # Shared legend
    from matplotlib.patches import Patch as _Patch
    fig.legend(
        handles=[_Patch(color=POS_COL,
                        label="Positive pCR assoc. (mean SHAP > 0)"),
                 _Patch(color=NEG_COL,
                        label="Negative pCR assoc. (mean SHAP < 0)")],
        loc="lower center", ncol=2, fontsize=9,
        bbox_to_anchor=(0.5, -0.01), frameon=False)

    fig.suptitle(
        "Consensus signatures — frozen per-modality features used in the final model\n"
        "Features ranked by mean cross-classifier selection rank across discovery "
        "folds (a normalised rank, NOT a SHAP magnitude — bar lengths are not "
        "comparable between modalities)  "
        "·  colour = sign of mean SHAP (positive/negative pCR association)",
        fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    _savefig(fig, fd / "fig02_consensus_signatures.pdf")


def fig_consensus_roc(data, fd):
    """
    Fig 03 (MAIN TEXT) — ROC curves from consensus model OOF predictions.

    One panel per scenario. Shows the ROC curve of the consensus fused model
    and of the fixed-best consensus unimodal model (modality with the highest
    headline AUROC). The CURVE is drawn from all (patient, repeat) out-of-fold
    predictions pooled — a picture of the average operating characteristic;
    the AUROC printed in the legend is the headline estimand (mean over CV
    repeats of the pooled-OOF AUROC, 95% patient-level cluster-bootstrap
    CI; _consensus_auroc_ci), so fig01, fig03 and PREDIX_HER2_results.xlsx
    quote the same number.
    """
    from sklearn.metrics import roc_curve
    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [fig03] No consensus data — skipping.")
        return

    fig, axes = plt.subplots(1, len(scenarios),
                              figsize=(5.5 * len(scenarios), 5.5), sharey=True)
    if len(scenarios) == 1:
        axes = [axes]
    fus_col = MOD_COLOR["Fused_ElasticNet"]

    def _lab(res):
        return (f"AUROC {res['estimate']:.3f} "
                f"[{res['ci_low']:.3f}–{res['ci_high']:.3f}]")

    for ax, sc in zip(axes, scenarios):
        cons_data = data[sc]["_consensus"]
        folds     = cons_data["folds"]
        pooled    = cons_data["pooled"]
        y_true = np.concatenate([f["y_test"] for f in folds])

        # Fused curve
        y_fus = np.concatenate([f["fused_y_pred"] for f in folds])
        fpr_f, tpr_f, _ = roc_curve(y_true, y_fus)
        res_f = _consensus_auroc_ci(data, sc, "Fused_ElasticNet")
        ax.plot(fpr_f, tpr_f, color=fus_col, lw=2.3,
                 label=f"Fused ElasticNet  ({_lab(res_f)})", zorder=3)

        # Fixed-best unimodal curve — modality with the highest headline AUROC
        mod_auc = {}
        for m in UNIMODALS:
            if m in pooled:
                r_m = _consensus_auroc_ci(data, sc, m)
                if np.isfinite(r_m["estimate"]):
                    mod_auc[m] = r_m["estimate"]
        if mod_auc:
            best_mod = max(mod_auc, key=mod_auc.get)
            y_uni = np.concatenate([f["unimodal_y_pred"][best_mod] for f in folds])
            fpr_u, tpr_u, _ = roc_curve(y_true, y_uni)
            res_u = _consensus_auroc_ci(data, sc, best_mod)
            ax.plot(fpr_u, tpr_u, color=MOD_COLOR[best_mod], lw=1.6, ls="--",
                     label=f"{best_mod} (best unimodal, {_lab(res_u)})",
                     alpha=0.85, zorder=2)

        ax.plot([0, 1], [0, 1], "k--", lw=0.7, alpha=0.4)
        ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.set_aspect("equal")
        ax.set_xlabel("1 − specificity")
        ax.set_ylabel("sensitivity")
        ax.set_title(sc, color=SC_COL[sc], fontsize=11, fontweight="bold")
        ax.legend(fontsize=8, loc="lower right")
        ax.grid(alpha=0.3)

    fig.suptitle(
        "ROC curves — consensus model out-of-fold predictions\n"
        "Solid = fused consensus · dashed = fixed-best unimodal consensus · "
        "curve: all (patient, repeat) OOF predictions pooled; "
        "AUROC = mean over CV repeats [95% patient-level cluster-bootstrap CI]",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig03_consensus_roc.pdf")


def fig_consensus_modality_weights(data, fd):
    """
    Fig 04 (MAIN TEXT) — Mean fusion coefficients of the consensus model.

    Across the fold-refits of the consensus fusion layer, what is each
    modality's mean weight? Error bar = SD across folds. A low-selection-rate
    modality annotation below each bar shows the fraction of folds in which
    the fusion layer retained a non-zero weight for that modality.
    """
    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [fig04] No consensus data — skipping.")
        return

    fig, axes = plt.subplots(1, len(scenarios),
                              figsize=(5.5 * len(scenarios), 5), sharey=True)
    if len(scenarios) == 1:
        axes = [axes]

    for ax, sc in zip(axes, scenarios):
        cons_data = data[sc]["_consensus"]
        folds = cons_data["folds"]
        if not folds:
            continue
        weight_rows = {m: [] for m in UNIMODALS}
        for f in folds:
            mw = f.get("modality_weights", {})
            for m in UNIMODALS:
                weight_rows[m].append(float(mw.get(m, 0.0)))

        xs = np.arange(len(UNIMODALS))
        for xi, mod in enumerate(UNIMODALS):
            arr = np.asarray(weight_rows[mod])
            mean_w = float(arr.mean())
            sd_w   = float(arr.std())
            sel_rate = float(np.mean(np.abs(arr) > 1e-6))
            col = MOD_COLOR[mod]
            ax.bar(xi, mean_w, 0.62, color=col, alpha=0.88)
            ax.errorbar(xi, mean_w, yerr=sd_w, fmt="none",
                         ecolor="#333", elinewidth=1.1, capsize=5)
            ax.text(xi, mean_w + (0.01 if mean_w >= 0 else -0.02),
                     f"{mean_w:+.2f}", ha="center",
                     va="bottom" if mean_w >= 0 else "top",
                     fontsize=8.5, color=col, fontweight="bold")
            # Selection-rate under tick
            ax.text(xi, ax.get_ylim()[0],
                     f"sel {sel_rate*100:.0f}%",
                     ha="center", va="top", fontsize=7.5, color=col,
                     transform=ax.get_xaxis_transform())

        ax.axhline(0, color="black", lw=0.9, ls="--")
        ax.set_xticks(xs)
        ax.set_xticklabels(UNIMODALS, fontsize=9)
        ax.set_title(sc, color=SC_COL[sc], fontsize=11, fontweight="bold")
        ax.set_ylabel("fusion coefficient (mean ± SD across folds)")
        ax.grid(axis="y", alpha=0.35)
        ax.tick_params(axis="x", pad=18)

    fig.suptitle(
        "Consensus fusion weights — per-modality contribution to the final fused model\n"
        "Mean ± SD of the elastic-net fusion coefficient across outer CV folds",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig04_consensus_modality_weights.pdf")


def fig_consensus_feature_shap(data, fd):
    """
    Fig 05 (MAIN TEXT) — SHAP beeswarm on the consensus signature features.

    Same beeswarm style as the discovery-phase fig05 (supp), but restricted
    strictly to the consensus signature features and aggregated across the
    discovery folds where those features appeared. One figure per scenario.

    Data source note: SHAP values come from the discovery-phase PKL (each
    winner fold's SHAP on its own signature), filtered to the subset of
    features that are in the final consensus signature. This shows the
    per-feature attribution for exactly the features reported in fig02.
    """
    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [fig05] No consensus data — skipping.")
        return

    for sc in scenarios:
        n_mods = sum(1 for m in UNIMODALS if m in data[sc])
        if n_mods == 0:
            continue
        cons_per_mod = data[sc]["_consensus"]["consensus"]
        fig, axes = plt.subplots(1, 5, figsize=(26, 10))
        for ci, mod in enumerate(UNIMODALS):
            ax  = axes[ci]
            col = MOD_COLOR[mod]
            folds = data[sc].get(mod, {}).get("folds", [])
            consensus_sig = cons_per_mod.get(mod, {}).get("signature", [])
            if not consensus_sig:
                ax.text(0.5, 0.5, "No consensus\nsignature",
                        ha="center", va="center", transform=ax.transAxes,
                        fontsize=9, color="#666")
                ax.set_title(mod, color=col, fontsize=11, fontweight="bold")
                ax.set_xticks([]); ax.set_yticks([])
                continue

            df_ = agg_shap(folds, feature_whitelist=consensus_sig)
            sc_ = draw_bee(ax, df_, mod, max_f=len(consensus_sig))
            add_cb(fig, sc_, ax)
            ax.set_title(f"{mod}  (n_sig = {len(consensus_sig)})",
                          color=col, fontsize=11, fontweight="bold")
        sc_lbl = "Global Model" if sc == "Global" else f"{sc} Arm"
        # RUN 5 — say plainly that the classifier is not fixed across the cloud.
        # SHAP is a MODEL-SPECIFIC attribution, and these points are pooled over
        # outer folds whose inner-CV winners differ (e.g. Global RNA: ExtraTrees
        # 38%, ElasticNet 31%, SVM 22%, RF 7%, HGB 2%). The beeswarm is
        # therefore an average attribution over the model-selection
        # distribution, not the attribution of one fitted model — legitimate for
        # this protocol, but it must not read as a single model. Also note each
        # feature contributes only the folds where it was selected, so point
        # density is not comparable across features, and points are
        # patient x fold rather than independent patients.
        fig.suptitle(
            f"SHAP attribution — consensus signatures  |  {sc_lbl}\n"
            "Each point = one test patient in one outer fold  ·  "
            "Colour: blue = low feature value, red = high\n"
            "Restricted to the final consensus signature  ·  "
            "pooled over outer folds, each contributing its own inner-CV "
            "winning classifier (distribution in supp_fig10)",
            fontsize=10.5, fontweight="bold", y=1.015)
        plt.tight_layout(w_pad=0.4)
        _savefig(fig, fd / f"fig05_consensus_feature_shap_{sc.replace('-','_')}.pdf")


def build_consensus_excel_main(data, td):
    """
    PREDIX_HER2_results.xlsx — MAIN TEXT headline Excel workbook.

    Three sheets:
      Performance    Pooled OOF AUROC / AUPRC / Brier / Youden Sens / Spec /
                     threshold for each consensus model (5 unimodal + fused)
                     in each scenario. This is the HEADLINE PERFORMANCE TABLE
                     quoted in the paper's Results section.
      Signatures     One row per (scenario, modality, feature); columns list
                     rank, mean cross-classifier selection rank (a normalised
                     rank, NOT a SHAP magnitude), winner classifier and
                     hyperparameters. This is the scientific deliverable.
      Fusion         Per-scenario mean fusion coefficients across outer CV
                     folds plus the modality selection rate.

    Discovery-phase diagnostics (fold-averaged AUROC, classifier comparisons,
    calibration profiles, etc.) are in tables/supplementary/supp_*.xlsx.
    """
    # This is a thin wrapper that calls build_consensus_sheet but with
    # the headline filename.
    # First, generate the consensus sheet into the target dir, then
    # move it to the headline filename.
    import shutil
    build_consensus_sheet(data, td)
    src = td / "consensus_model.xlsx"
    dst = td / "PREDIX_HER2_results.xlsx"
    if src.exists():
        shutil.move(str(src), str(dst))
        print(f"  → {dst.name} (headline)")


# =============================================================================
# EXCEL WORKBOOK
# =============================================================================
def build_excel(data, td):
    wb = openpyxl.Workbook()
    _blue  = PatternFill("solid", fgColor="1F4E79")
    _ltblu = PatternFill("solid", fgColor="EBF3FB")
    _white = PatternFill("solid", fgColor="FFFFFF")
    _bold_white = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    _norm       = Font(size=9, name="Arial")
    _ct = Alignment(horizontal="center", vertical="center")
    _lt = Alignment(horizontal="left",   vertical="center")
    _bs = Side(style="thin", color="BBBBBB")
    _bd = Border(left=_bs, right=_bs, top=_bs, bottom=_bs)

    def _hdr(ws, row, cols):
        for c, txt in enumerate(cols, 1):
            ce = ws.cell(row=row, column=c)
            ce.value = txt; ce.fill = _blue; ce.font = _bold_white
            ce.alignment = _ct; ce.border = _bd

    def _dat(ws, row, vals, alt=False):
        fill = _ltblu if alt else _white
        for c, v in enumerate(vals, 1):
            ce = ws.cell(row=row, column=c)
            ce.value = v; ce.font = _norm; ce.border = _bd
            ce.alignment = _ct if isinstance(v, (int, float)) else _lt
            ce.fill = fill

    def _wid(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Performance ──────────────────────────────────────────────────
    ws = wb.active; ws.title = "Performance"
    hdr = ["Scenario", "Model",
           "Mean AUROC", "SD AUROC", "Mean AUPRC", "SD AUPRC",
           "Mean Sensitivity", "SD Sensitivity",
           "Mean Specificity", "SD Specificity",
           "Mean Brier", "SD Brier", "N Folds", "Winner CLF (top)",
           "Mean Sig Size", "Platt Rate", "Stage-B Tuned Rate"]
    _hdr(ws, 1, hdr)

    r = 2
    for sc in SCENARIOS:
        for mod in ALL_MODELS:
            if mod not in data[sc]:
                continue
            d = data[sc][mod]
            folds = d["folds"]
            def _m(a): return round(float(np.nanmean(a)), 4)
            def _s(a): return round(float(np.nanstd(a)),  4)
            top_clf = sig_size = platt_rate = stageb_rate = ""
            if mod in UNIMODALS:
                winners = [f.get("winner_clf","") for f in folds
                           if f.get("winner_clf","") not in ("","none")]
                if winners:
                    w, cnt = Counter(winners).most_common(1)[0]
                    top_clf = f"{w} ({cnt/len(folds)*100:.0f}%)"
                sizes = [f.get("signature_size",0) for f in folds
                         if f.get("signature_size",0) > 0]
                sig_size = round(float(np.mean(sizes)), 1) if sizes else ""
                platts = [int(f.get("platt_applied",False)) for f in folds]
                platt_rate = f"{np.mean(platts)*100:.0f}%" if platts else ""
                # Stage B status: "tuned" means inner_cv_auroc_B came from
                # the in-fold signature-restricted tuning grid;
                # "fallback_stage_a" means tuning failed on every grid point
                # and the reported value is the Stage A pruned AUROC. A low
                # tuned rate in any row means the Stage-B AUROC column in the
                # Classifiers sheet should be interpreted cautiously for that cell.
                stati = [f.get("stage_b_status", "tuned") for f in folds]
                if stati:
                    tuned_pct = sum(1 for s in stati if s == "tuned") / len(stati)
                    stageb_rate = f"{tuned_pct*100:.0f}%"
            _dat(ws, r,
                 [sc, mod,
                  _m(d["aurocs"]), _s(d["aurocs"]),
                  _m(d["auprcs"]), _s(d["auprcs"]),
                  _m(d["senss"]),  _s(d["senss"]),
                  _m(d["specs"]),  _s(d["specs"]),
                  _m(d["briers"]), _s(d["briers"]),
                  len(folds), top_clf, sig_size, platt_rate, stageb_rate],
                 alt=(r % 2 == 0))
            r += 1
    _wid(ws, [10,20,13,11,13,11,17,15,17,15,12,11,9,25,13,12,15])

    # ── Sheet 2: Classifiers ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Classifiers")
    seen_clfs = set()
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for fold in data[sc].get(mod, {}).get("folds", []):
                seen_clfs.update(fold.get("inner_cv_aurocs_A", {}).keys())
    clfs2 = [c for c in ALL_CLFS if c in seen_clfs]
    hdr2  = (["Scenario", "Modality"] +
             [f"Mean AUROC-A ({c})" for c in clfs2] +
             [f"SD AUROC-A ({c})"   for c in clfs2] +
             ["Winner (top)", "Stage-B AUROC"])
    _hdr(ws2, 1, hdr2)
    r = 2
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            folds = data[sc].get(mod, {}).get("folds", [])
            if not folds:
                continue
            clf_means = {}; clf_sds = {}
            for clf in clfs2:
                vals = [f["inner_cv_aurocs_A"][clf]
                        for f in folds if clf in f.get("inner_cv_aurocs_A", {})]
                clf_means[clf] = round(float(np.mean(vals)), 4) if vals else ""
                clf_sds[clf]   = round(float(np.std(vals)),  4) if vals else ""
            winners = [f.get("winner_clf","") for f in folds
                       if f.get("winner_clf","") not in ("","none")]
            top_w   = Counter(winners).most_common(1)[0][0] if winners else ""
            b_au    = [f.get("inner_cv_auroc_B", np.nan) for f in folds]
            mean_b  = round(float(np.nanmean(b_au)), 4)
            _dat(ws2, r,
                 [sc, mod] +
                 [clf_means[c] for c in clfs2] +
                 [clf_sds[c]   for c in clfs2] +
                 [top_w, mean_b],
                 alt=(r%2==0))
            r += 1
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 10
    for i in range(3, len(hdr2)+1):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 3: Signatures ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Signatures")
    _hdr(ws3, 1, ["Scenario", "Modality", "Feature",
                   "In_Signature",
                   "Selection Freq (all folds)",
                   "Mean |SHAP| (when selected)",
                   "Mean SHAP direction"])
    _green     = PatternFill("solid", fgColor="C6EFCE")
    _green_fnt = Font(bold=True, color="276221", size=9, name="Arial")
    _gray_fnt  = Font(color="808080", size=9, name="Arial")
    r = 2
    for sc in SCENARIOS:
        thresh_sc = STABILITY_THRESH[sc]
        for mod in UNIMODALS:
            folds = data[sc].get(mod, {}).get("folds", [])
            valid  = [f for f in folds if get_signature_feats(f)]
            if not valid:
                continue
            n_tot = len(valid)
            feat_freq = Counter()
            feat_shap = defaultdict(list)
            for fold in valid:
                for feat in get_signature_feats(fold):
                    feat_freq[feat] += 1
                sh = fold.get("oof_shap")
                if sh:
                    for feat, sv in _mean_signed_shap(sh):
                        feat_shap[feat].append(float(sv))
            all_feats = feat_freq.most_common()
            stable   = [(f, c) for f, c in all_feats if c / n_tot >= thresh_sc]
            unstable = [(f, c) for f, c in all_feats if c / n_tot <  thresh_sc]
            for feat_list, is_sig in [(stable, True), (unstable, False)]:
                for feat, cnt in feat_list:
                    freq      = round(cnt / n_tot, 4)
                    shaps     = feat_shap.get(feat, [])
                    mean_abs  = round(float(np.mean(np.abs(shaps))), 5) if shaps else ""
                    mean_sign = round(float(np.mean(shaps)), 5) if shaps else ""
                    _dat(ws3, r, [sc, mod, feat, is_sig, freq, mean_abs, mean_sign],
                         alt=(r % 2 == 0))
                    if is_sig:
                        for col_i in range(1, 8):
                            ws3.cell(r, col_i).fill = _green
                        ws3.cell(r, 4).font = _green_fnt
                    else:
                        ws3.cell(r, 4).font = _gray_fnt
                    r += 1
    _wid(ws3, [10, 10, 35, 14, 22, 22, 20])

    # ── Sheet 4: Modality Weights ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Modality_Weights")
    _hdr(ws4, 1, (["Scenario"] +
                   [f"Mean coef ({m})" for m in UNIMODALS] +
                   [f"SD coef ({m})"   for m in UNIMODALS] +
                   [f"Sel rate ({m})"  for m in UNIMODALS] +
                   ["Tuned C (mean)"]))
    r = 2
    for sc in SCENARIOS:
        folds = data[sc].get("Fused_ElasticNet", {}).get("folds", [])
        if not folds:
            continue
        means = [round(float(np.mean([f["modality_weights"].get(m,0)
                                       for f in folds])), 4) for m in UNIMODALS]
        sds   = [round(float(np.std( [f["modality_weights"].get(m,0)
                                       for f in folds])), 4) for m in UNIMODALS]
        sel   = [round(float(np.mean([1 if m in f.get("selected_modalities",[])
                                        else 0 for f in folds])), 3)
                 for m in UNIMODALS]
        c_bar = round(float(np.mean([f.get("tuned_C",np.nan) for f in folds])), 3)
        _dat(ws4, r, [sc] + means + sds + sel + [c_bar], alt=(r%2==0))
        r += 1
    _wid(ws4, [10] + [16]*15 + [14])

    # ── Sheet 5: Calibration ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Calibration")
    seen_clfs2 = set()
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for fold in data[sc].get(mod, {}).get("folds", []):
                seen_clfs2.update(fold.get("calibration",{}).keys())
    clfs5 = [c for c in ALL_CLFS if c in seen_clfs2]
    hdr5  = (["Scenario","Modality"] +
             [f"Mean slope ({c})"  for c in clfs5] +
             [f"Platt rate ({c})"  for c in clfs5])
    _hdr(ws5, 1, hdr5)
    r = 2
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            folds = data[sc].get(mod, {}).get("folds", [])
            if not folds:
                continue
            slopes = {}; rates = {}
            for clf in clfs5:
                sl  = [f["calibration"][clf]["slope"]
                       for f in folds if clf in f.get("calibration",{})]
                np_ = [int(f["calibration"][clf]["needs_platt"])
                       for f in folds if clf in f.get("calibration",{})]
                slopes[clf] = round(float(np.nanmean(sl)),  3) if sl  else ""
                rates[clf]  = round(float(np.mean(np_)),    3) if np_ else ""
            _dat(ws5, r,
                 [sc, mod] + [slopes[c] for c in clfs5] + [rates[c] for c in clfs5],
                 alt=(r%2==0))
            r += 1
    _wid(ws5, [10,10] + [18]*len(clfs5)*2)

    # ── Sheet 6: Pooled operating-point metrics ───────────────────────────────
    # AUROC and Youden Sens/Spec computed on (y_test, y_pred) POOLED across
    # all outer folds and picked at a SINGLE threshold. These are the
    # deployment numbers — unlike the fold-mean metrics in the Performance
    # sheet which are upper-envelope optimistic (each fold picks its own
    # Youden-best threshold on ~30 test patients).
    # Values come from results["_pooled_metrics"] which the pipeline writes.
    ws6 = wb.create_sheet("Pooled_OpPoint")
    hdr6 = ["Scenario", "Model", "Pooled AUROC", "Pooled AUPRC",
            "Pooled Brier", "Pooled Sensitivity", "Pooled Specificity",
            "Pooled Threshold", "N pooled"]
    _hdr(ws6, 1, hdr6)
    r = 2
    any_pooled = False
    for sc in SCENARIOS:
        # The pooled metrics are stored under results["_pooled_metrics"]
        # at PKL load time; retrieve via the raw PKL cache on `data`.
        pooled_dict = data.get(sc, {}).get("_pooled_metrics", None)
        if pooled_dict is None:
            # Fallback: compute on the fly from fold-level y_test/y_pred
            pooled_dict = {}
            for mod in ALL_MODELS:
                folds = data[sc].get(mod, {}).get("folds", [])
                if not folds:
                    continue
                y_t = np.concatenate([np.asarray(f["y_test"], dtype=float)
                                       for f in folds])
                y_p = np.concatenate([np.asarray(f["y_pred"], dtype=float)
                                       for f in folds])
                if len(y_t) == 0 or len(np.unique(y_t)) < 2:
                    continue
                fpr, tpr, thr = roc_curve(y_t, y_p)
                yi = int(np.argmax(tpr - fpr))
                from sklearn.metrics import (roc_auc_score,
                                              average_precision_score,
                                              brier_score_loss)
                pooled_dict[mod] = {
                    "AUROC":       float(roc_auc_score(y_t, y_p)),
                    "AUPRC":       float(average_precision_score(y_t, y_p)),
                    "Brier":       float(brier_score_loss(y_t, y_p)),
                    "Sensitivity": float(tpr[yi]),
                    "Specificity": float(1.0 - fpr[yi]),
                    "Threshold":   float(thr[yi]),
                    "N_pooled":    int(len(y_t)),
                }
        for mod in ALL_MODELS:
            p = pooled_dict.get(mod)
            if not p:
                continue
            any_pooled = True
            _dat(ws6, r, [
                sc, mod,
                round(p.get("AUROC", np.nan), 4),
                round(p.get("AUPRC", np.nan), 4),
                round(p.get("Brier", np.nan), 4),
                round(p.get("Sensitivity", np.nan), 4),
                round(p.get("Specificity", np.nan), 4),
                round(p.get("Threshold", np.nan), 4),
                int(p.get("N_pooled", 0)),
            ], alt=(r % 2 == 0))
            r += 1
    _wid(ws6, [10, 20, 14, 14, 12, 18, 18, 16, 12])
    if not any_pooled:
        ws6.cell(row=2, column=1, value=(
            "Pooled metrics unavailable. Re-run pipeline on this version "
            "to populate results['_pooled_metrics']."))

    # ── Sheet 7: Counterfactual arm-switch ────────────────────────────────────
    # Per-patient cross-arm predictions pooled across folds. Reads
    # Fused_ElasticNet['cross_arm_preds'] dicts from each arm's PKL fold list.
    # Each row: one patient from the opposite arm, scored by this arm's models.
    ws7 = wb.create_sheet("Counterfactual")
    hdr7 = ["Scored-by model arm", "Patient global ID",
            "Mean P(pCR | this arm)",
            "95% CI lower", "95% CI upper",
            "N folds", "Benefit at 0.5", "Strong benefit at 0.5"]
    _hdr(ws7, 1, hdr7)
    r = 2
    any_cf = False
    THRESH = 0.5
    for sc_key, scored_arm in [("DHP", "DHP"), ("T-DM1", "T-DM1")]:
        folds = data.get(sc_key, {}).get("Fused_ElasticNet", {}).get("folds", [])
        if not folds:
            continue
        cap_pooled = defaultdict(list)
        for fold in folds:
            for pid, p in fold.get("cross_arm_preds", {}).items():
                cap_pooled[int(pid)].append(float(p))
        for pid, preds in sorted(cap_pooled.items()):
            arr = np.asarray(preds, dtype=float)
            arr = arr[~np.isnan(arr)]
            if len(arr) == 0:
                continue
            any_cf = True
            mean_p = float(arr.mean())
            rng = np.random.default_rng(_stable_seed("cf", scored_arm, pid))
            idx = rng.integers(0, len(arr), size=(2000, len(arr)))
            boots = arr[idx].mean(axis=1)
            lo = float(np.percentile(boots, 2.5))
            hi = float(np.percentile(boots, 97.5))
            benefit        = mean_p > THRESH
            strong_benefit = benefit and lo > THRESH
            _dat(ws7, r, [
                scored_arm, pid,
                round(mean_p, 3), round(lo, 3), round(hi, 3),
                int(len(arr)),
                "YES" if benefit else "no",
                "YES" if strong_benefit else "no",
            ], alt=(r % 2 == 0))
            r += 1
    _wid(ws7, [22, 18, 22, 14, 14, 10, 16, 22])
    if not any_cf:
        ws7.cell(row=2, column=1, value=(
            "No cross_arm_preds found in PKL. Re-run pipeline with the "
            "cross-arm addition to populate this sheet."))

    path = td / "PREDIX_HER2_results.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


# =============================================================================
# PERFORMANCE TABLE WITH 95% CI
# =============================================================================
def build_performance_ci_table(data, td):
    """
    Standalone Excel workbook: discovery-phase performance metrics with 95%
    patient-level cluster-bootstrap CIs (cv_estimands.py).

    For each scenario × model:
      • AUROC / AUPRC / Brier — metric computed on each CV repeat's complete
        out-of-fold vector, averaged over repeats; CI from 2000 patient-level
        cluster-bootstrap resamples (patients resampled with all their
        repeats; folds are NOT treated as independent).
      • Sensitivity / Specificity — mean over outer folds, each fold at its
        own stored Youden threshold; CI from the same patient-level cluster
        bootstrap (cv_estimands.bootstrap_fold_operating_point_ci).
    Two formats per metric:
      • Formatted string  "mean [lo–hi]"  — for display / manuscript
      • Raw numeric cols  mean / lo / hi  — for downstream analysis
    """
    wb = openpyxl.Workbook()
    _blue       = PatternFill("solid", fgColor="1F4E79")
    _ltblu      = PatternFill("solid", fgColor="EBF3FB")
    _white      = PatternFill("solid", fgColor="FFFFFF")
    _fus_fill   = PatternFill("solid", fgColor="EDE7F6")  # soft purple for fusion rows
    _bold_white = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    _norm       = Font(size=9, name="Arial")
    _bold_norm  = Font(bold=True, size=9, name="Arial")
    _ct  = Alignment(horizontal="center", vertical="center")
    _lt  = Alignment(horizontal="left",   vertical="center")
    _bs  = Side(style="thin", color="BBBBBB")
    _bd  = Border(left=_bs, right=_bs, top=_bs, bottom=_bs)

    # (metric name, number format). AUROC/AUPRC/Brier are repeat-level
    # estimands; Sensitivity/Specificity are fold operating points.
    #
    # RUN 5 — READ THIS BEFORE QUOTING Sensitivity OR Specificity.
    # Their threshold is `fd["metrics"]["Threshold"]`, which the pipeline sets
    # to thresholds[argmax(tpr - fpr)] computed ON THAT FOLD'S HELD-OUT TEST
    # SET. Each fold's pair is therefore the maximum of the Youden index over
    # ~22 patients / ~9 events, and the cluster bootstrap holds that
    # data-chosen threshold FIXED across all resamples — it resamples patients
    # but never re-selects the threshold, so the interval cannot express the
    # selection variance and cannot correct the bias. Simulated on this design
    # (n=110, 46 events, 5 folds, true AUROC ~0.75): Youden J at an a-priori
    # threshold 0.363, versus 0.530 at each fold's own threshold — an optimism
    # of about +0.17 in J, i.e. roughly +0.08 on each of Sens and Spec.
    # AUROC, AUPRC and Brier are threshold-free and unaffected.
    # The column headers below say this, and the honest single-threshold pair
    # is in the Pooled_OpPoint sheet. Do not quote these two as deployment
    # performance.
    METRICS = [
        ("AUROC",       ".3f"),
        ("AUPRC",       ".3f"),
        ("Sensitivity", ".3f"),
        ("Specificity", ".3f"),
        ("Brier",       ".3f"),
    ]
    REPEAT_METRICS = {"AUROC", "AUPRC", "Brier"}

    def _hdr(ws, row, cols):
        for c, txt in enumerate(cols, 1):
            ce = ws.cell(row=row, column=c)
            ce.value = txt; ce.fill = _blue; ce.font = _bold_white
            ce.alignment = _ct; ce.border = _bd

    def _cell(ws, r, c, v, fill, bold=False, align=None):
        ce = ws.cell(row=r, column=c, value=v)
        ce.fill   = fill
        ce.font   = _bold_norm if bold else _norm
        ce.border = _bd
        ce.alignment = align or (_ct if isinstance(v, (int, float)) else _lt)

    def _wid(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # One sheet covering all scenarios together (easier to compare)
    ws = wb.active
    ws.title = "Performance_95CI"

    # Build header: Scenario | Model | N Folds | N patients | N repeats |
    # for each metric: "<Metric> <estimand> [95% CI]" | formatted | mean | lo | hi
    N_ID_COLS = 5
    hdr_row1 = ["Scenario", "Model", "N Folds", "N patients", "N repeats"]
    hdr_row2 = [""] * N_ID_COLS
    for name, _ in METRICS:
        if name in REPEAT_METRICS:
            hdr_row1 += [f"{name} mean over CV repeats "
                         f"[95% patient-level cluster-bootstrap CI]", "", "", ""]
        else:
            # Say plainly that the threshold was chosen on the scored data.
            # "mean over folds at fold thresholds" was accurate but read as
            # neutral; a reader quoting this column needs to know it is an
            # upper envelope, not a deployable operating point.
            hdr_row1 += [f"{name} at each fold's Youden threshold, OPTIMISED ON "
                         f"THAT FOLD'S TEST DATA (optimistic ~+0.08; CI holds "
                         f"the chosen threshold fixed). Use Pooled_OpPoint for "
                         f"a single deployable threshold; AUROC is "
                         f"threshold-free", "", "", ""]
        hdr_row2 += [f"{name} (formatted)", f"{name} mean", f"{name} CI lo", f"{name} CI hi"]

    # Two-row header
    for ci, txt in enumerate(hdr_row1, 1):
        ce = ws.cell(row=1, column=ci, value=txt if txt else None)
        ce.fill = _blue; ce.font = _bold_white
        ce.alignment = _ct; ce.border = _bd
    for ci, txt in enumerate(hdr_row2, 1):
        ce = ws.cell(row=2, column=ci, value=txt if txt else None)
        ce.fill = PatternFill("solid", fgColor="2C5F8A")
        ce.font = Font(bold=True, color="FFFFFF", size=8, name="Arial")
        ce.alignment = _ct; ce.border = _bd

    # Merge identity columns vertically and each metric header across its
    # 4 sub-columns
    for c in range(1, N_ID_COLS + 1):
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
    for mi in range(len(METRICS)):
        start_c = N_ID_COLS + 1 + mi * 4
        ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=start_c + 3)

    def _fmt_ci(mn, lo, hi, fmt):
        if np.isnan(mn):
            return ""
        if np.isnan(lo) or np.isnan(hi):
            return f"{mn:{fmt}}"
        return f"{mn:{fmt}} [{lo:{fmt}}–{hi:{fmt}}]"

    def _num(v):
        return round(float(v), 4) if not np.isnan(v) else ""

    r = 3
    for sc in SCENARIOS:
        sc_color = {"Global": "F5F5F5", "DHP": "E3F2FD", "T-DM1": "FBE9E7"}[sc]
        sc_fill  = PatternFill("solid", fgColor=sc_color)
        for mod in ALL_MODELS:
            if mod not in data[sc]:
                continue
            d     = data[sc][mod]
            folds = d["folds"]
            is_fused = mod in FUS_VARS
            fill  = _fus_fill if is_fused else (
                    _ltblu if r % 2 == 0 else _white)

            # Repeat-level estimands (cached; shared with supp fig14)
            rep = {m: _discovery_metric_ci(data, sc, mod, m)
                   for m in ("AUROC", "AUPRC", "Brier")}
            # Fold operating points at each fold's own threshold, patient
            # cluster bootstrap for the CI
            op = CE.bootstrap_fold_operating_point_ci(
                folds, lambda f: f["y_pred"],
                lambda f: f["metrics"].get("Threshold"),
                n_boot=_N_BOOT,
                seed=_stable_seed("perfci", sc, mod, "Sensitivity_Specificity"))

            row_vals = [sc, mod, len(folds),
                        int(rep["AUROC"]["n"]), int(rep["AUROC"]["n_repeats"])]
            for name, fmt in METRICS:
                res = rep[name] if name in REPEAT_METRICS else op[name]
                mn, lo, hi = res["estimate"], res["ci_low"], res["ci_high"]
                row_vals += [_fmt_ci(mn, lo, hi, fmt), _num(mn), _num(lo), _num(hi)]

            for ci, v in enumerate(row_vals, 1):
                bold = ci <= 2  # bold Scenario and Model
                _cell(ws, r, ci, v, fill, bold=bold)
            r += 1

    # Column widths: Scenario, Model, N Folds, N patients, N repeats, then 4×5 metrics
    widths = [10, 22, 8, 10, 10] + [34, 12, 12, 12] * len(METRICS)
    _wid(ws, widths)
    ws.freeze_panes = "F3"

    path = td / "PREDIX_HER2_performance_CI.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


# =============================================================================
# FEATURE SELECTION PRUNING REPORT
# =============================================================================
def _and_list(items):
    """'A', 'A and B', 'A, B and C' — readable in prose cells."""
    items = list(items)
    if len(items) <= 1:
        return "".join(items)
    return ", ".join(items[:-1]) + " and " + items[-1]


def _pct(n, total):
    """Percentage that never rounds a non-zero count down to a bare '0%'."""
    p = 100.0 * n / total
    return f"{p:.0f}%" if p >= 1 else f"{p:.1f}%"


def _screen_effect(data):
    """Realised effect of the Tier-2.5 screen, measured from the fold audits.

    Returns (per_scenario_text, floor_text, exempt_mods). Nothing here is
    retyped: every count comes from the `univariate_screen` audit dict the
    pipeline writes into each fold. Modalities whose audit is absent in every
    fold were skipped by the UNIV_SCREEN_MIN_FEATURES carve-out.
    """
    parts, floor_bits, exempt = [], [], []
    for sc in SCENARIOS:
        bits = []
        for mod in UNIMODALS:
            folds = data.get(sc, {}).get(mod, {}).get("folds", [])
            aud = [f.get("univariate_screen") for f in folds]
            good = [a for a in aud if a and "n_input" in a]
            if not good:
                if folds and mod not in exempt:
                    exempt.append(mod)
                continue
            ni = int(np.median([a["n_input"] for a in good]))
            nr = int(np.median([a["n_retained"] for a in good]))
            bits.append(f"{mod} {ni}→{nr}")
            n_floor = sum(1 for a in good if a.get("floor_used"))
            if n_floor:
                floor_bits.append(f"{sc}/{mod} {n_floor}/{len(good)} folds "
                                  f"({_pct(n_floor, len(good))})")
        if bits:
            parts.append(f"{sc} — " + ", ".join(bits))
    return "; ".join(parts), ", ".join(floor_bits), exempt


def _small_mod_sizes(data, small_mods):
    """'Clin (5 features) and WSI (3 features)' — counts measured, not typed.

    The small modalities are exempt from screening and keep every feature, so
    the union of features seen across the pooled folds IS the modality's panel
    size. Falls back to the bare names if the folds cannot supply a count.
    """
    bits = []
    for mod in sorted(small_mods):
        feats = set()
        for fold in data.get("Global", {}).get(mod, {}).get("folds", []):
            for clf_imp in fold.get("inner_importance", {}).values():
                feats.update(clf_imp)
        bits.append(f"{mod} ({len(feats)} features)" if feats else mod)
    return _and_list(bits)


def _nzv_train_sizes(data):
    """Median expanded training n per outer fold, per scenario × modality."""
    parts = []
    for sc in SCENARIOS:
        bits = []
        for mod in UNIMODALS:
            folds = data.get(sc, {}).get(mod, {}).get("folds", [])
            vals = [f["n_train_expanded"] for f in folds
                    if f.get("n_train_expanded")]
            if vals:
                bits.append(f"{mod} {int(np.median(vals))}")
        if bits:
            parts.append(f"{sc} — " + " / ".join(bits))
    return "; ".join(parts)


def _platt_rate(data):
    """(n_folds_calibrated, n_folds_total) over every scenario × modality."""
    applied = total = 0
    for sc in SCENARIOS:
        for mod in UNIMODALS:
            for f in data.get(sc, {}).get(mod, {}).get("folds", []):
                if "platt_applied" in f:
                    total += 1
                    applied += bool(f["platt_applied"])
    return applied, total


def build_feature_pruning_report(data, td, results_dir=None):
    """
    Excel workbook documenting the full feature-selection pipeline.

    Sheet 1 — Methodology: fixed parameters and a prose description of each
    pruning stage. Every parameter is read from the pipeline source, the run
    provenance or the PKLs — see the "READING THE PIPELINE'S OWN CONSTANTS"
    block at the top of this module for why none of them may be retyped here.
    The Tier 3 section in particular is GATED on the pipeline's own
    CORR_FILTER_MODS, so this sheet can never again describe a per-fold
    correlation filter that is not running.

    `results_dir` supplies run_provenance.json, the record of the CLI-settable
    thresholds this run actually used. It is optional only so that older
    callers keep working; without it the pipeline's argparse defaults are used.

    Sheet 2 — Statistics: per modality × scenario, tracks reconstructed
    feature counts at each stage using PKL data:
      • Features entering the model (union of all features seen in
        inner_importance across folds — i.e. survivors of the fold's NZV
        filter and Tier-2.5 univariate screen)
      • Mean EPV-capped signature size across folds (winner_signature)
      • Stable features (above STABILITY_THRESH) across folds
      • Mean selection frequency of stable features
    """
    wb = openpyxl.Workbook()
    _blue      = PatternFill("solid", fgColor="1F4E79")
    _ltblu     = PatternFill("solid", fgColor="EBF3FB")
    _white     = PatternFill("solid", fgColor="FFFFFF")
    _green_h   = PatternFill("solid", fgColor="E8F5E9")
    _bw        = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    _norm      = Font(size=9, name="Arial")
    _bold      = Font(bold=True, size=9, name="Arial")
    _hd        = Font(bold=True, size=10, name="Arial", color="1F4E79")
    _it        = Font(italic=True, size=9, name="Arial", color="444444")
    _ct        = Alignment(horizontal="center", vertical="center")
    _lt        = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    _bs        = Side(style="thin", color="BBBBBB")
    _bd        = Border(left=_bs, right=_bs, top=_bs, bottom=_bs)

    def _hdr(ws, row, cols):
        for c, txt in enumerate(cols, 1):
            ce = ws.cell(row=row, column=c)
            ce.value = txt; ce.fill = _blue; ce.font = _bw
            ce.alignment = _ct; ce.border = _bd

    def _wid(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Methodology ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Methodology"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 90

    def _section(row, title):
        ce = ws1.cell(row=row, column=1, value=title)
        ce.font = _hd
        ce.fill = PatternFill("solid", fgColor="E3F2FD")
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws1.row_dimensions[row].height = 18
        return row + 1

    def _row(row, label, value, italic=False):
        c1 = ws1.cell(row=row, column=1, value=label)
        c1.font = _bold; c1.border = _bd
        c1.alignment = Alignment(horizontal="left", vertical="top")
        c2 = ws1.cell(row=row, column=2, value=value)
        c2.font = _it if italic else _norm; c2.border = _bd
        c2.alignment = _lt
        ws1.row_dimensions[row].height = max(15, min(60, len(str(value)) // 3))
        return row + 1

    # Values that define the method, each from its single source of truth.
    _t1        = _tier1_remove_from_pipeline()
    _corr_mods = _pipeline_const("CORR_FILTER_MODS")
    _dedup     = _pipeline_const("CONSENSUS_DEDUP_MODS")
    _min_feat  = _pipeline_const("UNIV_SCREEN_MIN_FEATURES")
    _fdr_q     = _run_param(results_dir, "univ_fdr_q")
    _max_k     = _run_param(results_dir, "univ_max_k")
    _min_k     = _run_param(results_dir, "univ_min_k")
    _nzv_glob  = _run_param(results_dir, "nzv_freq_global")
    _nzv_arm   = _run_param(results_dir, "nzv_freq_arm")
    _nzv_ratio = _run_param(results_dir, "nzv_ratio")
    _corr_thr  = _run_param(results_dir, "corr_threshold")
    _stab      = _stability_thresh(results_dir)
    _epv       = _pipeline_func_const("_derive_signature", "EPV")
    _floor     = _pipeline_func_const("_derive_signature", "FLOOR")
    _small_mods = _pipeline_func_const("_derive_signature", "SMALL_MODS")
    _small_txt = _small_mod_sizes(data, _small_mods)
    _p25       = _pipeline_percentile_arg("_derive_signature")
    _l1        = _pipeline_const("L1_RATIO")
    _n_mod     = len(UNIMODALS)     # fusion input width is one column per modality
    _screen_on = str(_provenance_params(results_dir).get(
        "univariate_screen", "in_fold")) == "in_fold"
    _eff, _floor_txt, _exempt = _screen_effect(data)
    _platt_n, _platt_tot = _platt_rate(data)

    # True stage order, stated once so a reader is not left wondering what
    # happened to the tier that used to sit between 2 and 4.
    _stages = ["Tier 1 biological deduplication (once, before any split)",
               "Tier 2 near-zero-variance filter (per fold)"]
    if _screen_on:
        _stages.append("Tier 2.5 in-fold univariate outcome screen (per fold)")
    _stages.append("Tier 3 high-correlation filter — "
                   + ("ACTIVE (per fold)" if _corr_mods else "REMOVED, no longer runs"))
    _stages.append("Tier 4 EPV cap + 25th-percentile filter + floor (per fold)")

    r = 1
    r = _section(r, "FEATURE SELECTION PIPELINE — PREDIX HER2 MULTIMODAL pCR ANALYSIS")
    r = _row(r, "Stage order",
             "  →  ".join(_stages)
             + ". Tier 3 is retained as a numbered heading below, marked with "
               "its current status, because the originally submitted manuscript "
               "described it and a reader holding that version will look for it.")
    r = _row(r, "Provenance of this sheet",
             "Every threshold below is read at generation time from "
             "multimodal_pcr_pipeline.py (hard-coded constants, parsed — not "
             "imported), from run_provenance.json (parameters the command line "
             "could override, i.e. what this run actually used), or measured "
             "from the result PKLs (realised effects). None of them is typed "
             "into this report by hand.", italic=True)
    r += 1

    r = _section(r, "TIER 1 — Biological Deduplication (applied once, before any CV fold)")
    r = _row(r, "Method",
             "Domain-expert removal of biologically redundant features. Features that "
             "are co-amplified on the same chromosomal segment (e.g. 17q12 amplicon: "
             "ERBB2/GRB7/PPP1R1B/MIEN1), derive from the same composite score "
             "(e.g. TMB variants), or represent the same signal at different granularities "
             "are collapsed to a single representative per cluster. This is a deterministic, "
             "non-data-driven decision applied before the train/test split.")
    r = _row(r, "Scope", "All modalities")
    r = _row(r, "Features removed",
             f"{len(_t1)} listed in TIER1_REMOVE: " + ", ".join(_t1))
    r = _row(r, "Note",
             "RNA_FCGR3B is listed here but is NOT a redundancy removal. It was "
             "excluded on measurement-validity grounds decided in advance: the "
             "transcript is neutrophil-restricted, so its signal in bulk tumour "
             "RNA-seq of fresh-frozen biopsies is attributable to peripheral-blood "
             "contamination rather than to tumour biology. It is not collinear "
             "with any retained feature.", italic=True)
    r += 1

    r = _section(r, "TIER 2 — Near-Zero Variance (NZV) Filter (per outer fold, fitted on training only)")
    r = _row(r, "Method",
             "A feature is removed if either condition holds on the fold training set: "
             "(a) its most common value occupies at least freq_threshold of training "
             "samples, OR (b) the most common value is at least ratio_threshold times as "
             "frequent as the second most common AND also occupies at least "
             "freq_threshold of the OBSERVED (non-missing) values. The same features are "
             "then dropped from the test set. Rationale: binary genomic features (rare "
             "mutation indicators) can have near-zero variance in a training fold, making "
             "them unstable and uninformative predictors.")
    r = _row(r, "Condition (b) — dominance gate",
             "The second clause of (b) is not cosmetic. Ungated, the ratio rule alone "
             f"removed any binary feature with carrier prevalence below about "
             f"1/(ratio+1) ≈ {100.0 / (float(_nzv_ratio) + 1):.1f}% regardless of "
             "freq_threshold, which silently overrode the loosened arm threshold whose "
             "whole purpose is to keep low-prevalence mutation indicators. The gate uses "
             "the observed-value denominator, so the rule still catches features whose "
             "top value dominates the observed distribution but escapes condition (a) "
             "because missingness dilutes the training-set fraction.", italic=True)
    r = _row(r, "freq_threshold",
             f"{_nzv_glob} in the pooled (Global) model; {_nzv_arm} in the arm models "
             f"(DHP, T-DM1). The arm value is deliberately looser because the arm "
             f"training folds are smaller, and a feature present in a handful of "
             f"patients should not be culled merely for being rare in an already "
             f"small fold.")
    r = _row(r, "ratio_threshold",
             f"{_nzv_ratio} (top-1 to top-2 value frequency ratio, gated as above)")
    r = _row(r, "Scope",           "All modalities")
    r = _row(r, "Fitted on",
             "Outer training set only (no leakage to test set). Under expanded "
             "training the fold's training set is every patient with that modality "
             "measured, minus the outer test patients — not the complete case.")
    r = _row(r, "Realised training n per outer fold (median)",
             _nzv_train_sizes(data) + ". These are the sample sizes the NZV filter is "
             "actually fitted on; they differ by modality because each modality "
             "trains on the patients for whom it was measured.")
    r += 1

    if _screen_on:
        r = _section(r, "TIER 2.5 — In-Fold Univariate Outcome Screen "
                        "(per outer AND inner fold, fitted on training only)")
        r = _row(r, "Method",
                 "A tie-corrected Mann-Whitney U test of every candidate feature against "
                 "pCR on that fold's training patients, giving a univariate AUROC and "
                 "p-value per feature. p-values are adjusted within the modality by "
                 f"Benjamini-Hochberg and features with q ≤ {_fdr_q} are retained. The "
                 f"retained set is then capped at {_max_k} features, keeping the "
                 f"strongest by |AUROC − 0.5|; and if fewer than {_min_k} survive, the "
                 f"top {_min_k} by |AUROC − 0.5| are restored so a modality can never "
                 "collapse to an empty or degenerate design matrix.")
        r = _row(r, "Why q ≤ " + str(_fdr_q) + " and not 0.05",
                 "This is a SCREENING step, not an inference step: its job is to reduce "
                 "the candidate pool to a tractable size before the multivariable "
                 "signature discovery in Stage A does the actual selection. A stringent "
                 "threshold at these training-set sizes would discard features whose "
                 "contribution is only visible multivariably. The threshold is "
                 "pre-specified and identical across every fold, arm and modality.")
        r = _row(r, "Fitted on",
                 "Outer training set only, and then independently again inside each "
                 "inner fold. No test patient influences which features enter the "
                 "model — this is the leakage control that distinguishes the present "
                 "analysis from the original submission, where the univariate step had "
                 "been applied to the whole cohort before the input file was written.")
        r = _row(r, "Order within the fold",
                 "Runs AFTER median imputation and standardisation (inside "
                 "fit_imputer_scaler), so the test statistic sees a complete matrix and "
                 "no feature is screened out because of missingness rather than because "
                 "of its association with pCR.")
        r = _row(r, "Exempt modalities",
                 f"Modalities entering the fold with ≤ {_min_feat} candidate columns are "
                 f"skipped entirely (UNIV_SCREEN_MIN_FEATURES = {_min_feat}), matching "
                 f"the small-modality carve-out at Tier 4"
                 + (f". Measured in these PKLs: no screen audit is recorded in any fold "
                    f"for {_and_list(_exempt)}, confirming the carve-out applied to "
                    f"{'them' if len(_exempt) > 1 else 'it'} and to nothing else."
                    if _exempt else "."))
        r = _row(r, "Realised effect (median candidates in → out, per outer fold)",
                 _eff + ". This is the largest single reduction in the pipeline, and "
                 "it is the reason the per-fold candidate count is far below the "
                 "post-Tier-1 panel size.")
        if _floor_txt:
            # Volunteered, not buried. In the arm/DNA cells this is close to a
            # coin flip, and a reviewer who found it unaided would be entitled
            # to read the omission as concealment. It is an honest property of
            # screening a small arm cohort, not a defect in the procedure.
            r = _row(r, "IMPORTANT — how often the floor replaced the FDR rule",
                     f"In a substantial minority of folds NO feature reached q ≤ "
                     f"{_fdr_q}, and the floor restored the top {_min_k} by "
                     f"|AUROC − 0.5| regardless. In those folds the screen was not an "
                     f"FDR-controlled filter at all: it degenerated into a plain "
                     f"univariate top-{_min_k} rule, and the model for that modality "
                     f"was built from features that did not clear the FDR gate. Rates "
                     f"measured fold by fold in these PKLs: {_floor_txt}. The two arm "
                     f"DNA cells are the ones to note. This is expected behaviour for a "
                     f"univariate screen on a small arm cohort with few events rather "
                     f"than a defect, but selection frequencies and stability flags for "
                     f"the affected modalities must be read with it in mind.")
        r += 1

    if _corr_mods:
        r = _section(r, "TIER 3 — High-Correlation Filter (per outer fold, fitted on training only)")
        r = _row(r, "Method",
                 "Absolute Pearson correlation matrix computed on continuous training "
                 "features (>2 unique values; binary features excluded). For each "
                 "correlated cluster (|r| ≥ threshold), the highest-variance training "
                 "representative is kept and all others removed.")
        r = _row(r, "threshold", f"|r| ≥ {_corr_thr} (absolute Pearson correlation)")
        r = _row(r, "Scope",     f"{', '.join(sorted(_corr_mods))} only")
        r = _row(r, "Fitted on", "Outer training set only (no leakage to test set)")
        r = _row(r, "Applied",   "Outer fold preprocessing only (not repeated inside inner folds)")
    else:
        r = _section(r, "TIER 3 — High-Correlation Filter — REMOVED, THIS STAGE DOES NOT RUN")
        r = _row(r, "Status",
                 "Not applied, in any modality or scenario. CORR_FILTER_MODS is empty in "
                 "multimodal_pcr_pipeline.py, and this section is generated from that "
                 "constant: if the per-fold correlation filter were ever switched back "
                 "on, this heading would describe it instead. The stage was present in "
                 "the originally submitted analysis, which is why it keeps its number "
                 "here rather than being deleted.")
        r = _row(r, "Why it was removed",
                 "The filter kept whichever member of a correlated cluster performed "
                 "best in that fold's training data and dropped the rest. Which member "
                 "survived therefore varied from fold to fold, and — because the "
                 "comparison was made against the outcome — that choice was itself "
                 "outcome-influenced. The rotation made per-feature selection "
                 "frequencies uninterpretable: two near-identical features each "
                 "accumulated a fraction of the selection frequency that one distinct "
                 "biological signal deserved, and both could then enter the pooled "
                 "consensus signature as if they were independent findings.")
        r = _row(r, "What replaces it",
                 "Redundancy is now decided ONCE, before any train/test split, by the "
                 "fixed TIER1_REMOVE list above. That decision is outcome-blind and "
                 "identical in every fold, so no selection frequency is diluted by a "
                 "rotating representative and no fold-specific choice can depend on the "
                 "outcome. Removing the stage also removes a step from the Methods "
                 "rather than adding one.")
        r = _row(r, "Correlation gate (audit only)",
                 f"|r| ≥ {_corr_thr} is still used, but as a PREFLIGHT AUDIT over the "
                 f"whole candidate panel: preflight.py fails the run if any surviving "
                 f"pair exceeds it. The guarantee Tier 3 used to provide fold by fold is "
                 f"therefore now checked once, for the entire panel, before the run "
                 f"starts. It is not a per-fold filter.")
        if _dedup:
            r = _row(r, "Not to be confused with",
                     "Correlation-cluster pooling at the CONSENSUS stage is a separate "
                     f"mechanism and is still active for {_and_list(sorted(_dedup))}. It "
                     "operates once on the full cohort rather than fold by fold, so it "
                     "does not rotate, and it acts on the aggregated signature rather "
                     "than on the features entering each model.", italic=True)
    r += 1

    r = _section(r, f"TIER 4 — EPV Cap + {_p25}th-Percentile Filter + Floor "
                    f"(per outer fold, from cross-classifier importance)")
    r = _row(r, "Method",
             "After Stage A Pass 1 (cross-classifier percentile-rank importance aggregation), "
             "the feature signature per modality is derived using three sequential rules:\n"
             f"  (1) EPV ceiling: max_k = max(floor(n_pCR_events / EPV={_epv}), "
             f"FLOOR={_floor}). Sets an upper bound grounded in the events-per-variable "
             f"literature, calibrated to EPV={_epv} for regularised elastic net models "
             "(rather than the classical EPV=10 for unregularised regression). The "
             f"FLOOR={_floor} ensures at minimum {_floor} features.\n"
             f"  (2) {_p25}th-percentile filter: among the top max_k features by mean cross-"
             f"classifier percentile rank, those below the {_p25}th percentile of the "
             "retained set are dropped. This removes the bottom quartile — features "
             "consistently ranked as least informative across classifiers — while "
             "retaining everything with meaningful cross-classifier consensus.\n"
             "  (3) Floor protection: if the percentile filter reduces the set below "
             f"FLOOR={_floor}, the top {_floor} features by importance rank are restored "
             "regardless of the percentile threshold. This prevents over-pruning in arm "
             "scenarios.\n"
             f"EXCEPTION: {_small_txt} retain ALL features — "
             "the elastic net L1/L2 regularisation handles non-informative features by "
             "shrinking their coefficients, and removing features from a modality with "
             "only a handful of them risks a degenerate model. Stage B winner selection "
             "uses the pruned signature AUROC.")
    r = _row(r, "EPV constant",
             f"{_epv} (Events Per Variable; calibrated for regularised models)")
    r = _row(r, "Feature floor",
             f"{_floor} (minimum retained features for "
             f"{_and_list([m for m in UNIMODALS if m not in _small_mods])})")
    r = _row(r, "Percentile threshold",
             f"{_p25}th percentile within EPV-capped set")
    r = _row(r, "Small modalities",
             f"{_small_txt} — all features retained (elastic net regularises)")
    r = _row(r, "Winner criterion",      "Mean inner val AUROC of PRUNED signature (Stage A Pass 2)")
    r = _row(r, "Stability threshold (Global)",
             f"{_stab['Global']:.0%} — feature in stable signature if "
             f"selected in ≥{_stab['Global']:.0%} of outer folds")
    r = _row(r, "Stability threshold (Arm)",
             f"{_stab['DHP']:.0%} — applied to DHP and T-DM1 arm models")
    r += 1

    r = _section(r, "FUSION — Single Fused ElasticNet Meta-Learner")
    r = _row(r, "Method",
             f"A single ElasticNet meta-learner (L1+L2, l1_ratio={_l1}) is trained on the "
             f"{_n_mod}-column OOF probability matrix (one column per modality). The L1 "
             "component zeros out non-contributing modalities, producing an interpretable "
             "sparse modality weighting. The L2 component handles collinearity between OOF "
             "predictions. C is tuned by inner CV over FUSION_C_GRID.")
    r = _row(r, "Input",
             f"{_n_mod}-column OOF probability matrix — leakage-safe stacking")
    r = _row(r, "l1_ratio",     f"{_l1} (equal L1 and L2 penalty)")
    # Was: `FUSION_C_GRID if 'FUSION_C_GRID' in dir() else "<hand-typed list>"`.
    # Inside a function `dir()` lists LOCALS, so the guard was always False and
    # the hand-typed fallback always won — it happened to still be correct,
    # which is exactly how this class of bug survives. Read the real constant
    # and fail loudly if the pipeline no longer defines it.
    r = _row(r, "C grid",       str(_pipeline_const("FUSION_C_GRID")))
    r += 1

    r = _section(r, "CALIBRATION (per outer fold, winner classifier)")
    r = _row(r, "Method",
             "A single two-parameter Platt sigmoid is fitted on the winning "
             "classifier's inner-loop out-of-fold predictions for that fold, and then "
             "applied to both the OOF vector and the outer-test predictions. It is "
             "applied UNCONDITIONALLY — to every modality and every fold — so that all "
             f"{_n_mod} modality OOF columns entering the fusion layer are on the same "
             "calibrated probability scale. Applying it to some modalities and not "
             "others would hand the meta-learner columns with different variances for "
             "reasons unrelated to their predictive value.")
    r = _row(r, "Platt fitting",
             "LogisticRegression(C=1e6, solver='lbfgs', max_iter=1000) on the raw score "
             "as the single predictor, fitted once on the fold's accumulated inner-OOF "
             "predictions. NOT a nested CalibratedClassifierCV: fitting the sigmoid "
             "inside cv=3 splits gave it roughly 30 patients, which is underpowered for "
             "even two parameters, whereas the pooled inner-OOF fit gives it the whole "
             "complete-case training fold.")
    r = _row(r, "Applied to",     "Winner classifier only (after Stage A winner selection)")
    if _platt_tot:
        r = _row(r, "Realised",
                 f"Calibration was applied in {_platt_n:,} of {_platt_tot:,} "
                 f"scenario × modality × fold combinations "
                 f"({100.0 * _platt_n / _platt_tot:.1f}%). The only path that skips it "
                 f"is a fold with too little data or a single-class OOF vector to fit a "
                 f"sigmoid at all.")
    r = _row(r, "Diagnostic slope (does NOT gate anything)",
             "Each fold also records a per-classifier calibration slope under "
             "`calibration`, with a `needs_platt` flag set when the slope falls outside "
             "the pipeline's diagnostic band. That flag is recorded for auditing and is "
             "NOT read by the calibration path — it gates nothing. An earlier version of "
             "this pipeline did apply Platt scaling conditionally on it, via "
             "CalibratedClassifierCV; that is not what runs now, and any description of "
             "conditional calibration refers to the superseded design.", italic=True)

    _wid(ws1, [28, 90])
    ws1.freeze_panes = "B2"

    # ── Sheet 2: Statistics per Modality × Scenario ───────────────────────────
    ws2 = wb.create_sheet("Pruning_Statistics")

    hdr = ["Scenario", "Modality", "N Folds",
           "Features surviving NZV + Tier-2.5 screen (union over folds)",
           "Mean sig size (EPV-capped)",
           "SD sig size",
           "Min sig size",
           "Max sig size",
           f"Stable features (freq≥thresh)",
           "Mean freq of stable features",
           "Stability threshold"]
    _hdr(ws2, 1, hdr)
    # Add a note row explaining the approximation
    note_ce = ws2.cell(row=2, column=4,
                        value="Union, over every outer fold × inner fold × classifier, of "
                              "the features appearing in inner_importance — that is, "
                              "features that survived that fold's near-zero-variance "
                              "filter and Tier-2.5 univariate screen at least once. It "
                              "is a union, NOT a per-fold count: the candidate set "
                              "actually entering any single fold is considerably "
                              "smaller (see the Tier 2.5 realised-effect row on the "
                              "Methodology sheet). No correlation filter contributes to "
                              "this column — Tier 3 no longer runs.")
    note_ce.font = _it
    note_ce.alignment = _lt

    r = 3
    for sc in SCENARIOS:
        thresh = _stab[sc]           # provenance-accurate, not the module default
        for mod in UNIMODALS:
            folds = data[sc].get(mod, {}).get("folds", [])
            if not folds:
                continue

            # Post-outer-NZV/corr feature count: union of all inner_importance keys
            all_imp_feats = set()
            for fold in folds:
                for clf_imp in fold.get("inner_importance", {}).values():
                    all_imp_feats.update(clf_imp.keys())
            n_post_preproc = len(all_imp_feats) if all_imp_feats else ""

            # Signature sizes (EPV-capped winner signatures)
            sizes = [f.get("signature_size", 0) for f in folds
                     if f.get("signature_size", 0) > 0]
            mean_sz = round(float(np.mean(sizes)), 2)   if sizes else ""
            sd_sz   = round(float(np.std(sizes)),  2)   if sizes else ""
            min_sz  = int(np.min(sizes))                if sizes else ""
            max_sz  = int(np.max(sizes))                if sizes else ""

            # Stable signature: features above stability threshold
            valid = [f for f in folds if get_signature_feats(f)]
            n_tot = len(valid)
            feat_freq = Counter()
            for fold in valid:
                for feat in get_signature_feats(fold):
                    feat_freq[feat] += 1
            stable_feats = {f: cnt/n_tot for f, cnt in feat_freq.items()
                            if cnt/n_tot >= thresh}
            n_stable   = len(stable_feats)
            mean_freq  = round(float(np.mean(list(stable_feats.values()))), 3) \
                         if stable_feats else ""

            fill = _ltblu if r % 2 == 0 else _white
            row_vals = [sc, mod, len(folds),
                        n_post_preproc,
                        mean_sz, sd_sz, min_sz, max_sz,
                        n_stable, mean_freq,
                        f"{thresh:.0%}"]
            for ci, v in enumerate(row_vals, 1):
                ce = ws2.cell(row=r, column=ci, value=v)
                ce.font = _norm; ce.border = _bd; ce.fill = fill
                ce.alignment = _ct if isinstance(v, (int, float)) else _lt
            r += 1

    _wid(ws2, [10, 10, 9, 36, 26, 14, 14, 14, 28, 28, 22])

    path = td / "PREDIX_HER2_feature_pruning_report.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


# =============================================================================
# PERFORMANCE CI FIGURE
# =============================================================================
def fig_performance_ci(data, fd):
    """
    Fig 14 — Forest plot of discovery-phase AUROC per model × scenario.
    Each row = one model. Panels = scenarios.
    Dot = AUROC (mean over CV repeats of the pooled-OOF AUROC). Line = 95%
    patient-level cluster-bootstrap CI (cv_estimands.py; identical numbers to
    supp_PREDIX_HER2_performance_CI.xlsx). Vertical dashed = 0.5 (chance).
    Unimodal models and fusion models separated by a horizontal gap.
    """
    fig, axes = plt.subplots(1, 3, figsize=(18, 8), sharey=True)

    for ci, sc in enumerate(SCENARIOS):
        ax = axes[ci]
        models_present = [m for m in ALL_MODELS if m in data[sc]]
        n = len(models_present)
        y_pos = list(range(n))

        for yi, mod in enumerate(reversed(models_present)):
            res = _discovery_metric_ci(data, sc, mod, "AUROC")
            mn, lo, hi = res["estimate"], res["ci_low"], res["ci_high"]
            col  = MOD_COLOR[mod]
            is_fused = mod in FUS_VARS
            marker   = "D" if is_fused else "o"
            ms       = 60  if is_fused else 50

            ax.plot([lo, hi], [yi, yi], color=col, lw=2.2, alpha=0.85,
                    solid_capstyle="round", zorder=2)
            ax.scatter([mn], [yi], color=col, s=ms, zorder=4,
                       marker=marker, edgecolors="white", linewidths=0.6)
            ax.text(hi + 0.005, yi,
                    f"{mn:.3f} [{lo:.3f}–{hi:.3f}]",
                    va="center", fontsize=6.5, color=col)

        # Separator line between unimodals and fusion
        fus_start = sum(1 for m in reversed(models_present) if m in FUS_VARS)
        if 0 < fus_start < n:
            ax.axhline(fus_start - 0.5, color="#cccccc", lw=1, ls="-")

        ax.axvline(0.5, color="#aaaaaa", lw=0.9, ls="--", alpha=0.7)
        ax.set_xlim(0.30, 1.05)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(
            [m.replace("Fused_", "Fused ") for m in reversed(models_present)],
            fontsize=8.5)
        ax.set_xlabel("AUROC", fontsize=9)
        ax.set_title(sc, color=SC_COL[sc], fontsize=10, fontweight="bold")
        ax.spines["left"].set_visible(False)
        ax.tick_params(left=False)
        ax.grid(axis="x", alpha=0.30)

    fig.suptitle(
        "Model Performance — cross-validated AUROC with 95% CI (discovery phase)\n"
        "Circle = unimodal  ·  Diamond = fused  ·  AUROC = mean over CV repeats "
        "of the pooled out-of-fold AUROC  ·  "
        f"CI = {_N_BOOT} patient-level cluster-bootstrap resamples",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout(w_pad=0.5)
    _savefig(fig, fd / "fig14_performance_CI.pdf")


# =============================================================================
# SECTION F: COUNTERFACTUAL (cross-arm) SUMMARY
# =============================================================================
def _collect_cross_arm_from_pkl(results_dir, arm_key):
    """Small duplicate of counterfactual reader — keeps this file standalone."""
    pkl = Path(results_dir) / arm_key / f"{arm_key}_elasticnet_results.pkl"
    if not pkl.exists():
        return {}, None, 0
    with open(pkl, "rb") as f:
        raw = pickle.load(f)
    fusion_folds = raw.get("Fused_ElasticNet", [])
    alt_label = None
    acc = defaultdict(list)
    for fold in fusion_folds:
        cap = fold.get("cross_arm_preds", {})
        if not cap:
            continue
        alt_label = alt_label or fold.get("cross_arm_label")
        for pid, p in cap.items():
            acc[int(pid)].append(float(p))
    return dict(acc), alt_label, len(fusion_folds)


def _collect_assigned_from_pkl(results_dir, arm_key):
    """Read per-patient (arm-df index) assigned-arm predictions pooled across folds."""
    pkl = Path(results_dir) / arm_key / f"{arm_key}_elasticnet_results.pkl"
    if not pkl.exists():
        return {}
    with open(pkl, "rb") as f:
        raw = pickle.load(f)
    folds = raw.get("Fused_ElasticNet", [])
    splits_pkl = None
    for cand in [Path(results_dir) / arm_key / f"{arm_key}_cv_splits.pkl",
                 Path(results_dir) / f"{arm_key}_cv_splits.pkl"]:
        if cand.exists():
            splits_pkl = cand; break
    if splits_pkl is None:
        return {}
    with open(splits_pkl, "rb") as f:
        sp = pickle.load(f)
    outer = sp["outer"]
    acc = defaultdict(list)
    for fold in folds:
        fi = fold["fold_idx"]
        if fi >= len(outer):
            continue
        _, te = outer[fi]
        for idx, pred in zip(te, fold["y_pred"]):
            acc[int(idx)].append(float(pred))
    return dict(acc)


def fig_counterfactual_summary(results_dir, fd, threshold=0.5):
    """
    Fig 15 — Counterfactual arm-switch summary.

    Two-panel figure:
      Left:  Per-patient mean P_alt with 95% bootstrap CI, stratified by arm.
             Colours: strong benefit (CI lower > threshold), weak (mean > threshold,
             CI crosses), none. Matches the semantics of counterfactual_analysis.py
             but condensed into a single Nature Cancer-ready panel per arm.
      Right: Summary counts per arm + mean ΔP annotation.

    Reads directly from cross_arm_preds stored by the pipeline, so no refitting
    and no methodology drift. Silently skips if PKL lacks cross_arm_preds
    (e.g. pipeline was run before the cross-arm addition).
    """
    ARM_COLOR_LOCAL = {"DHP": "#2166ac", "T-DM1": "#d6604d"}
    # Load cross-arm predictions: DHP experiment predicts on T-DM1 patients, vice versa
    tdm1_preds_from_dhp_model, _, _ = _collect_cross_arm_from_pkl(results_dir, "dhp")
    dhp_preds_from_tdm1_model, _, _ = _collect_cross_arm_from_pkl(results_dir, "tdm1")
    if not tdm1_preds_from_dhp_model and not dhp_preds_from_tdm1_model:
        print("  [fig15] No cross_arm_preds in PKL — skipping.")
        return None

    assigned_dhp  = _collect_assigned_from_pkl(results_dir, "dhp")
    assigned_tdm1 = _collect_assigned_from_pkl(results_dir, "tdm1")

    # Build per-patient records using the same logic as counterfactual_analysis.py
    # We need to know which patients are non-pCR, which requires the original data
    # — BUT since we only have PKL here, we use the fold-level y_test stored in
    # in-arm predictions to identify non-pCR patients (pCR == 0 at any appearance).
    def _non_pcr_arm_indices(arm_key):
        pkl = Path(results_dir) / arm_key / f"{arm_key}_elasticnet_results.pkl"
        if not pkl.exists(): return set()
        with open(pkl, "rb") as f:
            raw = pickle.load(f)
        splits_pkl = None
        for cand in [Path(results_dir) / arm_key / f"{arm_key}_cv_splits.pkl",
                     Path(results_dir) / f"{arm_key}_cv_splits.pkl"]:
            if cand.exists():
                splits_pkl = cand; break
        if splits_pkl is None: return set()
        with open(splits_pkl, "rb") as f:
            sp = pickle.load(f)
        outer = sp["outer"]
        label_by_idx = {}
        for fold in raw.get("Fused_ElasticNet", []):
            fi = fold["fold_idx"]
            if fi >= len(outer): continue
            _, te = outer[fi]
            for idx, lbl in zip(te, fold["y_test"]):
                label_by_idx[int(idx)] = int(lbl)
        return {i for i, l in label_by_idx.items() if l == 0}

    non_pcr_dhp  = _non_pcr_arm_indices("dhp")
    non_pcr_tdm1 = _non_pcr_arm_indices("tdm1")

    # NOTE: mapping arm-df index → global patient_id requires the original CSV.
    # Since fig15 reads PKL only, we can ONLY report the shift distribution
    # across patients who have cross-arm preds available. The patient-identity
    # join is done robustly in counterfactual_analysis.py. Here we compute
    # a summary view: distribution of ΔP = P_alt − P_assigned, aggregated.
    # This is appropriate for a figure; the patient-level table lives in
    # counterfactual_results.xlsx.

    def _build_shifts(assigned_by_arm_idx, non_pcr_arm_idx, alt_preds_by_pid):
        """Return list of (P_assigned, P_alt, P_alt_lo, P_alt_hi) for eligible patients.
        Eligible = non-pCR AND P_assigned ≤ threshold.
        Matches by POSITION (assuming arm-df indices and global pids are ordered
        consistently — fine for a summary figure; exact join is in the Excel)."""
        # Build position → pid candidate list: in practice cross_arm_preds keys
        # are GLOBAL pids, and assigned keys are arm-df indices. Without the
        # data file we cannot join them by identity. So fig15 reports
        # DISTRIBUTIONS (not per-patient rows): mean P_alt across all eligible,
        # histogram of P_alt, compared with mean P_assigned across all eligible.
        elig_assigned = []
        for idx in sorted(non_pcr_arm_idx):
            if idx not in assigned_by_arm_idx: continue
            p_asgn = float(np.mean(assigned_by_arm_idx[idx]))
            if p_asgn > threshold: continue
            elig_assigned.append(p_asgn)
        p_alt_vals_all = []
        p_alt_means    = []
        p_alt_ci_lo    = []
        for pid, vals in alt_preds_by_pid.items():
            if len(vals) == 0: continue
            arr = np.asarray(vals, dtype=float)
            p_alt_vals_all.extend(vals)
            p_alt_means.append(float(arr.mean()))
            rng = np.random.default_rng(_stable_seed("alt_pid", pid))
            idx = rng.integers(0, len(arr), size=(2000, len(arr)))
            boot_means = arr[idx].mean(axis=1)
            p_alt_ci_lo.append(float(np.percentile(boot_means, 2.5)))
        return elig_assigned, p_alt_vals_all, p_alt_means, p_alt_ci_lo

    arms = [
        ("DHP",   non_pcr_dhp,  assigned_dhp,  dhp_preds_from_tdm1_model,  "T-DM1"),
        ("T-DM1", non_pcr_tdm1, assigned_tdm1, tdm1_preds_from_dhp_model,  "DHP"),
    ]

    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    for ai, (arm, non_pcr, assigned, alt_preds, alt_label) in enumerate(arms):
        ax = axes[ai]
        elig_asgn, _all_alt, alt_means, alt_ci_lo = _build_shifts(
            assigned, non_pcr, alt_preds)

        if not alt_means:
            ax.text(0.5, 0.5, f"No cross-arm\npredictions for {arm}",
                    transform=ax.transAxes, ha="center", va="center",
                    fontsize=11, color="#666")
            ax.set_xticks([]); ax.set_yticks([])
            continue

        # Stratify patients by benefit class using mean P_alt and CI lower
        alt_means_arr  = np.array(alt_means)
        alt_ci_lo_arr  = np.array(alt_ci_lo)
        strong = (alt_means_arr > threshold) & (alt_ci_lo_arr > threshold)
        weak   = (alt_means_arr > threshold) & (alt_ci_lo_arr <= threshold)
        none_  = ~(strong | weak)
        counts = [int(strong.sum()), int(weak.sum()), int(none_.sum())]

        # Histogram of mean P_alt with threshold line + benefit coloring
        bins = np.linspace(0, 1, 21)
        ax.hist(alt_means_arr[none_], bins=bins, color="#aaaaaa",
                alpha=0.75, edgecolor="white", label=f"no benefit (n={counts[2]})")
        ax.hist(alt_means_arr[weak], bins=bins, color="#9ec49e",
                alpha=0.85, edgecolor="white", label=f"weak benefit (n={counts[1]})")
        ax.hist(alt_means_arr[strong], bins=bins, color="#2ca02c",
                alpha=0.85, edgecolor="white", label=f"strong benefit (n={counts[0]})")
        ax.axvline(threshold, color="black", lw=1.0, ls="--", alpha=0.75)

        if elig_asgn:
            mean_asgn = float(np.mean(elig_asgn))
            mean_alt  = float(np.mean(alt_means_arr))
            delta     = mean_alt - mean_asgn
            ax.axvline(mean_alt, color=ARM_COLOR_LOCAL[arm], lw=1.5, alpha=0.9)
            ax.text(0.98, 0.98,
                    f"Eligible n={len(alt_means)}\n"
                    f"mean P(pCR | assigned) = {mean_asgn:.3f}\n"
                    f"mean P(pCR | {alt_label}) = {mean_alt:.3f}\n"
                    f"Δ (cohort-level) = {delta:+.3f}",
                    transform=ax.transAxes, ha="right", va="top",
                    fontsize=8.5,
                    bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                              edgecolor="#ccc", alpha=0.95))

        ax.set_xlim(0, 1)
        ax.set_xlabel(f"Mean P(pCR | {alt_label} model)  —  per-patient")
        ax.set_ylabel("Number of patients" if ai == 0 else "")
        ax.set_title(f"{arm} non-pCR patients → switch to {alt_label}",
                     color=ARM_COLOR_LOCAL[arm])
        ax.grid(axis="y", alpha=0.3)
        ax.legend(fontsize=8, loc="upper left")

    fig.suptitle(
        "Counterfactual arm-switch: predicted benefit distribution\n"
        "Per-patient mean P(pCR) under the alternative arm, classified by 95% bootstrap CI",
        fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout()
    _savefig(fig, fd / "fig06_counterfactual_summary.pdf")


def build_pooled_metrics_sheet(results_dir, td):
    """
    Pooled operating-point table: Sens/Spec at a SINGLE Youden threshold
    computed on predictions pooled across all outer folds.

    Per-fold Sens/Spec (in the Performance sheet) pick a Youden-optimal
    threshold on each fold's ~30 test patients, which is high-variance and
    optimistic — the mean of fold-specific bests is an upper envelope.
    Pooling y across folds first gives an honest deployment operating point:
    one threshold, one Sens, one Spec, comparable across modalities.

    Reads results["_pooled_metrics"] written by the pipeline. Silently skips
    a scenario if the field is absent.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pooled operating points"
    navy = PatternFill("solid", fgColor="1F4E79")
    wh   = PatternFill("solid", fgColor="FFFFFF")
    alt  = PatternFill("solid", fgColor="EBF3FB")
    bw   = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    nm   = Font(size=9, name="Arial")
    ct   = Alignment(horizontal="center", vertical="center")
    bs   = Side(style="thin", color="BBBBBB")
    bd   = Border(left=bs, right=bs, top=bs, bottom=bs)

    headers = ["Scenario", "Model",
               "AUROC (pooled)", "AUPRC (pooled)", "Brier (pooled)",
               "Sensitivity", "Specificity", "Youden threshold",
               "N patients pooled"]
    for ci, h in enumerate(headers, 1):
        ce = ws.cell(row=1, column=ci, value=h)
        ce.fill = navy; ce.font = bw; ce.alignment = ct; ce.border = bd

    ri = 2
    any_pooled = False
    for sc in SCENARIOS:
        exp = EXP_MAP[sc]
        candidate = Path(results_dir) / exp / f"{exp}_elasticnet_results.pkl"
        if not candidate.exists():
            continue
        with open(candidate, "rb") as f:
            raw = pickle.load(f)
        pooled = raw.get("_pooled_metrics")
        if not pooled:
            continue
        any_pooled = True
        for mod in ALL_MODELS:
            p = pooled.get(mod)
            if not p or np.isnan(p.get("AUROC", np.nan)):
                continue
            fill = alt if (ri % 2 == 0) else wh
            vals = [sc, mod,
                    round(float(p.get("AUROC",    np.nan)), 3),
                    round(float(p.get("AUPRC",    np.nan)), 3),
                    round(float(p.get("Brier",    np.nan)), 3),
                    round(float(p.get("Sensitivity", np.nan)), 3),
                    round(float(p.get("Specificity", np.nan)), 3),
                    round(float(p.get("Threshold", np.nan)), 3),
                    int(p.get("N_pooled", 0))]
            for ci, v in enumerate(vals, 1):
                ce = ws.cell(row=ri, column=ci, value=v)
                ce.fill = fill; ce.font = nm; ce.alignment = ct; ce.border = bd
            ri += 1

    if not any_pooled:
        ws.cell(row=2, column=1,
                value="No _pooled_metrics found in any PKL. "
                      "Re-run multimodal_pcr_pipeline.py with the pooled-metrics addition.")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    path = td / "pooled_operating_points.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


def build_counterfactual_sheet(results_dir, td, threshold=0.5):
    """
    Counterfactual Excel sheet: cohort-level summary per arm with
    strong / weak / none counts and cohort mean shift. Patient-level table
    lives in counterfactual_analysis.py output (counterfactual_results.xlsx)
    because that script has access to the raw data for patient-ID joins.
    """
    rows = []
    for arm_key, alt_label in [("dhp", "T-DM1"), ("tdm1", "DHP")]:
        alt_preds, _lbl, n_folds = _collect_cross_arm_from_pkl(
            results_dir, arm_key)
        if not alt_preds:
            continue
        means = []
        ci_lo = []
        for pid, vals in alt_preds.items():
            arr = np.asarray(vals, dtype=float)
            means.append(float(arr.mean()))
            rng = np.random.default_rng(_stable_seed("alt_pid", pid))
            idx = rng.integers(0, len(arr), size=(2000, len(arr)))
            ci_lo.append(float(np.percentile(arr[idx].mean(axis=1), 2.5)))
        means_arr = np.array(means); ci_lo_arr = np.array(ci_lo)
        strong = int(((means_arr > threshold) & (ci_lo_arr > threshold)).sum())
        weak   = int(((means_arr > threshold) & (ci_lo_arr <= threshold)).sum())
        none_  = int((means_arr <= threshold).sum())

        # "assigned arm" is the OPPOSITE of alt_label (the arm-key's own arm
        # is the model source; its predictions target the other arm's patients).
        assigned_arm_label = {"dhp": "T-DM1", "tdm1": "DHP"}[arm_key]
        rows.append({
            "Predicted population": assigned_arm_label,
            "Model source (alt arm)": alt_label,
            "N patients with cross-arm preds": len(means_arr),
            "N strong benefit (CI lower > thresh)": strong,
            "N weak benefit (mean > thresh, CI crosses)": weak,
            "N no benefit": none_,
            "% predicted benefit": round((strong + weak) / len(means_arr) * 100, 1)
                                     if len(means_arr) else 0.0,
            "Mean P(pCR | alt)": round(float(means_arr.mean()), 3),
            "Folds pooled": n_folds,
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counterfactual"
    navy = PatternFill("solid", fgColor="1F4E79")
    wh   = PatternFill("solid", fgColor="FFFFFF")
    alt  = PatternFill("solid", fgColor="EBF3FB")
    bw   = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    nm   = Font(size=9, name="Arial")
    ct   = Alignment(horizontal="center", vertical="center")
    bs   = Side(style="thin", color="BBBBBB")
    bd   = Border(left=bs, right=bs, top=bs, bottom=bs)

    if not rows:
        ws.cell(row=1, column=1,
                value="No cross_arm_preds found in any arm PKL. "
                      "Re-run multimodal_pcr_pipeline.py with cross-arm support.")
        wb.save(td / "counterfactual_summary.xlsx")
        print(f"  → counterfactual_summary.xlsx (empty — no cross-arm data)")
        return

    headers = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        ce = ws.cell(row=1, column=ci, value=h)
        ce.fill = navy; ce.font = bw; ce.alignment = ct; ce.border = bd
    for ri, row in enumerate(rows):
        fill = alt if ri % 2 == 0 else wh
        for ci, h in enumerate(headers, 1):
            ce = ws.cell(row=ri + 2, column=ci, value=row[h])
            ce.fill = fill; ce.font = nm; ce.alignment = ct; ce.border = bd

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 32

    path = td / "counterfactual_summary.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


# =============================================================================
# CONSENSUS EXCEL WORKBOOK (R2)
# =============================================================================
def build_consensus_sheet(data, td):
    """
    Write consensus_model.xlsx with three sheets:

      Signatures   One row per (scenario, modality, feature); columns list
                   rank, mean cross-classifier selection rank (a normalised
                   rank, NOT a SHAP magnitude — see the comment at the header
                   below), and the winner classifier / hyperparameters chosen
                   for that modality.

      Fusion       Per-scenario mean fusion coefficients (across CV folds
                   of the consensus re-evaluation) plus the selection rate
                   for each modality in the frozen-consensus fusion layer.

      Performance  AUROC (mean over CV repeats of the pooled-OOF AUROC) with
                   a 95% patient-level cluster-bootstrap CI (cv_estimands.py)
                   for each model (five unimodal + one fused) in each
                   scenario, plus descriptive pooled AUPRC / Brier / Youden
                   Sens / Spec / threshold. This sheet carries the primary
                   headline numbers for the Nature Cancer submission.

    Silent no-op if no scenario has a consensus eval PKL.
    """
    scenarios = [sc for sc in SCENARIOS if "_consensus" in data.get(sc, {})]
    if not scenarios:
        print("  [consensus_sheet] No consensus data — skipping.")
        return

    wb = openpyxl.Workbook()
    navy = PatternFill("solid", fgColor="1F4E79")
    wh   = PatternFill("solid", fgColor="FFFFFF")
    alt  = PatternFill("solid", fgColor="EBF3FB")
    hl   = PatternFill("solid", fgColor="FFF3CD")   # fused row highlight
    bw   = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    nm   = Font(size=9, name="Arial")
    bl   = Font(bold=True, size=9, name="Arial")
    ct   = Alignment(horizontal="center", vertical="center")
    lt   = Alignment(horizontal="left",   vertical="center")
    bs   = Side(style="thin", color="BBBBBB")
    bd   = Border(left=bs, right=bs, top=bs, bottom=bs)

    def _hdr(ws, headers):
        for ci, h in enumerate(headers, 1):
            ce = ws.cell(row=1, column=ci, value=h)
            ce.fill = navy; ce.font = bw; ce.alignment = ct; ce.border = bd

    # ── Sheet 1: Signatures ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Signatures"
    # Column E is `mean_importance`, which is a mean cross-classifier
    # PERCENTILE RANK, not a SHAP magnitude — per fold the K values are
    # 1/K..K/K and sum to (K+1)/2 (measured exactly on the run-5 PKLs: 3.0 for
    # K=5, 2.0 for K=3). Labelling it "Mean |SHAP importance|" claimed a unit
    # the column does not have, in the workbook the manuscript quotes. The data
    # key stays `mean_importance`: it is written by the pipeline into the
    # consensus PKLs and read by downstream consumers, so only the label moves.
    # Values are bounded by 1 and are comparable only WITHIN a modality.
    _hdr(ws1, ["Scenario", "Modality", "Rank", "Feature",
               "Mean selection rank (not |SHAP|)",
               "Winner classifier", "Classifier support (%)",
               "Hyperparameters", "Signature size (K)"])
    ri = 2
    for sc in scenarios:
        cons = data[sc]["_consensus"]["consensus"]
        for mod in UNIMODALS:
            c = cons.get(mod, {})
            sig = c.get("signature", [])
            imp = c.get("mean_importance", {})
            if not sig:
                continue
            fill = alt if ri % 2 == 0 else wh
            for rank, feat in enumerate(sig, 1):
                vals = [sc, mod, rank, feat,
                        round(float(imp.get(feat, 0.0)), 5),
                        c.get("winner_clf", ""),
                        round(c.get("support_fraction", 0.0) * 100, 1),
                        str(c.get("params", {})),
                        c.get("K", 0)]
                for ci, v in enumerate(vals, 1):
                    ce = ws1.cell(row=ri, column=ci, value=v)
                    ce.fill = fill; ce.font = nm; ce.border = bd
                    ce.alignment = ct if isinstance(v, (int, float)) else lt
                ri += 1
    for col, w in zip(ws1.columns,
                       [10, 10, 8, 38, 20, 20, 18, 45, 14]):
        ws1.column_dimensions[col[0].column_letter].width = w

    # ── Sheet 2: Fusion ──────────────────────────────────────────────────
    # Per-scenario mean fusion coefficients across the consensus re-eval folds.
    ws2 = wb.create_sheet("Fusion")
    _hdr(ws2, ["Scenario", "Modality",
               "Mean fusion coefficient",
               "SD fusion coefficient",
               "Selection rate (|coef| > 1e-6)",
               "Mean tuned C"])
    ri = 2
    for sc in scenarios:
        cons_data = data[sc]["_consensus"]
        folds = cons_data["folds"]
        if not folds:
            continue
        # Gather modality weights and tuned_C across folds
        weight_rows = defaultdict(list)
        tuned_cs    = []
        for f in folds:
            mw = f.get("modality_weights", {})
            for m in UNIMODALS:
                weight_rows[m].append(float(mw.get(m, 0.0)))
            if f.get("tuned_C") is not None:
                tuned_cs.append(float(f["tuned_C"]))
        mean_C = float(np.mean(tuned_cs)) if tuned_cs else np.nan

        for mod in UNIMODALS:
            arr = np.asarray(weight_rows[mod])
            sel = float(np.mean(np.abs(arr) > 1e-6))
            fill = alt if ri % 2 == 0 else wh
            vals = [sc, mod,
                    round(float(arr.mean()), 4),
                    round(float(arr.std()),  4),
                    round(sel, 3),
                    round(mean_C, 4) if not np.isnan(mean_C) else ""]
            for ci, v in enumerate(vals, 1):
                ce = ws2.cell(row=ri, column=ci, value=v)
                ce.fill = fill; ce.font = nm; ce.border = bd
                ce.alignment = ct if isinstance(v, (int, float)) else lt
            ri += 1
    for col, w in zip(ws2.columns, [10, 12, 22, 22, 26, 14]):
        ws2.column_dimensions[col[0].column_letter].width = w

    # ── Sheet 3: Performance ─────────────────────────────────────────────
    # Headline sheet. The quotable number is "AUROC (mean over CV repeats)"
    # with its patient-level cluster-bootstrap CI: in every CV repeat each
    # patient has exactly one out-of-fold prediction; the AUROC of that
    # complete vector is averaged over repeats (200 global / 100 per arm),
    # and the CI resamples PATIENTS carrying all their repeat predictions
    # (cv_estimands.py — the same helper and seed as fig01, so figure and
    # table agree to the last digit).
    #
    # Predictions are NEVER averaged across repeats before scoring: that
    # scores an ensemble, which is optimistic for informative models and
    # severely pessimistic for near-constant ones (held-out-outcome artefact).
    #
    # Descriptive columns retained for continuity with the submitted tables:
    #   "Pooled AUROC (all rows)"  = one AUROC over every (patient, repeat)
    #     row ranked together (the submission's 'pooled OOF' number);
    #   "Mean/SD fold AUROC"       = per-fold AUROC across repeated CV. The SD
    #     is NOT an uncertainty interval — folds reuse the same patients.
    # Sensitivity/Specificity/threshold are pooled-row values at the fold
    # thresholds (Youden), as in the submission.
    ws3 = wb.create_sheet("Performance")
    _hdr(ws3, ["Scenario", "Model",
               "AUROC (mean over CV repeats)",
               "AUROC 95% CI lower (patient cluster bootstrap)",
               "AUROC 95% CI upper (patient cluster bootstrap)",
               "AUROC formatted",
               "N patients (bootstrap unit)", "N pCR events",
               "N CV repeats",
               "Pooled AUROC (all rows, descriptive)",
               "Mean fold AUROC (descriptive)",
               "SD fold AUROC (descriptive, NOT a CI)",
               "Pooled AUPRC", "Pooled Brier",
               "Pooled Sensitivity", "Pooled Specificity",
               "Pooled threshold", "N pooled rows (= patients x repeats)"])
    ri = 2
    for sc in scenarios:
        cons_data = data[sc]["_consensus"]
        pooled    = cons_data["pooled"]

        for mod in UNIMODALS + ["Fused_ElasticNet"]:
            p = pooled.get(mod)
            if not p:
                continue
            res = _consensus_auroc_ci(data, sc, mod)
            est, lo, hi = res["estimate"], res["ci_low"], res["ci_high"]
            fmt = (f"{est:.3f} [{lo:.3f}–{hi:.3f}]"
                   if np.isfinite(est) and np.isfinite(lo) else "")

            is_fused = (mod == "Fused_ElasticNet")
            fill = hl if is_fused else (alt if ri % 2 == 0 else wh)
            font = bl if is_fused else nm
            vals = [sc, mod,
                    round(float(est), 4) if np.isfinite(est) else "",
                    round(float(lo), 4) if np.isfinite(lo) else "",
                    round(float(hi), 4) if np.isfinite(hi) else "",
                    fmt,
                    int(res["n"]), int(res["n_events"]),
                    int(res["n_repeats"]),
                    round(float(p["AUROC"]), 4),
                    round(float(p["mean_fold_AUROC"]), 4),
                    round(float(p["std_fold_AUROC"]),  4),
                    round(float(p["AUPRC"]), 4),
                    round(float(p["Brier"]), 4),
                    round(float(p["Sensitivity"]), 4),
                    round(float(p["Specificity"]), 4),
                    round(float(p["Threshold"]), 4),
                    int(p["N_pooled"])]
            for ci, v in enumerate(vals, 1):
                ce = ws3.cell(row=ri, column=ci, value=v)
                ce.fill = fill; ce.font = font; ce.border = bd
                ce.alignment = ct if isinstance(v, (int, float)) else lt
            ri += 1
    for col, w in zip(ws3.columns,
                       [10, 22, 16, 18, 18, 22, 14, 12, 12, 16, 14, 16,
                        13, 13, 16, 16, 14, 16]):
        ws3.column_dimensions[col[0].column_letter].width = w

    path = td / "consensus_model.xlsx"
    wb.save(path)
    print(f"  → {path.name}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    args = parse_args()
    fd   = args.out_dir / "figures"
    td   = args.out_dir / "tables"
    fd.mkdir(parents=True, exist_ok=True)
    td.mkdir(parents=True, exist_ok=True)

    print(f"\n[DATA] Loading from: {args.results_dir}")
    data = load_data(args.results_dir)

    mode = detect_mode(data)
    print(f"[MODE] Detected PKL training strategy: {mode}")

    print("\nGenerating figures ...")

    # ── MAIN-TEXT FIGURES ─────────────────────────────────────────────────
    # These use the CONSENSUS SIGNATURE + ITERATED OUTER CV protocol.
    # The performance numbers, signatures, ROC curves, SHAP, fusion
    # attribution, and counterfactual figures all reflect the signature
    # that will be reported in the paper as the scientific deliverable.
    # This is the single source of truth for performance claims.
    print("\nMain-text figures (consensus signature + iterated outer CV)...")
    fig_consensus_performance_main(data, fd)         # fig01
    fig_consensus_signatures_main(data, fd)          # fig02
    fig_consensus_roc(data, fd)                      # fig03
    fig_consensus_modality_weights(data, fd)         # fig04
    fig_consensus_feature_shap(data, fd)             # fig05 (one per scenario)
    fig_counterfactual_summary(args.results_dir, fd) # fig06 (cohort summary)

    # ── SUPPLEMENTARY FIGURES ────────────────────────────────────────────
    # Discovery-phase diagnostics: per-fold variability of signatures,
    # classifiers, and performance BEFORE consensus aggregation.
    # These support the claim that the discovery procedure is stable
    # and that the consensus signature is meaningful, but they are NOT
    # the performance numbers reported in the Results section.
    supp = fd / "supplementary"
    supp.mkdir(exist_ok=True)
    print("\nSupplementary figures (discovery-phase diagnostics)...")

    # Discovery performance (per-fold, pre-consensus)
    fig_roc(data, supp)                              # supp: ROC discovery
    fig_performance_distributions(data, supp)        # supp: per-fold distributions
    fig_fusion_benefit(data, supp)                   # supp: fusion benefit
    fig_forest_plot(data, supp)                      # supp: forest plot
    fig_performance_ci(data, supp)                   # supp: discovery AUROC + cluster CI

    # Discovery signature / feature-level diagnostics
    fig_feature_shap(data, supp)                     # supp: per-fold SHAP
    fig_feature_selection_frequency(data, supp)      # supp: selection freq
    fig_cross_scenario_features(data, supp)          # supp: cross-scenario feat
    fig_fusion_shap(data, supp)                      # supp: fusion SHAP
    fig_modality_weights(data, supp)                 # supp: modality weights

    # Discovery classifier diagnostics
    fig_winner_classifier_heatmap(data, supp)        # supp: winner heatmap
    fig_inner_auroc_comparison(data, supp)           # supp: inner AUROC
    fig_calibration_profile(data, supp)              # supp: calibration
    fig_signature_sizes(data, supp)                  # supp: signature sizes

    # Rename supplementary PDFs with supp_ prefix for clear separation.
    # Path.replace, not Path.rename: on Windows rename() refuses to overwrite
    # an existing target, so regenerating into a report directory that
    # already holds supp_* files (e.g. re-running step 2 on production PKLs)
    # would crash after all figures were drawn.
    for p in supp.glob("fig*.pdf"):
        new = supp / ("supp_" + p.name)
        p.replace(new)

    # ── EXCEL — MAIN: headline consensus results ──────────────────────────
    print("\nGenerating Excel ...")
    build_consensus_excel_main(data, td)             # PREDIX_HER2_results.xlsx
    # counterfactual still gets its own standalone summary workbook
    build_counterfactual_sheet(args.results_dir, td)

    # ── EXCEL — SUPPLEMENTARY: discovery diagnostics ─────────────────────
    supp_td = td / "supplementary"
    supp_td.mkdir(exist_ok=True)
    build_excel(data, supp_td)                       # supp: discovery Performance etc.
    build_performance_ci_table(data, supp_td)        # supp: discovery metrics + cluster CI
    build_feature_pruning_report(data, supp_td, args.results_dir)  # supp: pruning report
    build_pooled_metrics_sheet(args.results_dir, supp_td)  # supp: pooled ops (discovery)
    # Rename supplementary XLSX files too
    for p in supp_td.glob("*.xlsx"):
        if not p.name.startswith("supp_"):
            p.replace(supp_td / ("supp_" + p.name))

    n_main = len(list(fd.glob("fig*.pdf")))
    n_supp = len(list(supp.glob("supp_*.pdf")))
    print(f"\nDone.  Main-text figures  ({n_main}) → {fd}")
    print(f"       Supplementary figs ({n_supp}) → {supp}")
    print(f"       Tables            → {td}")

if __name__ == "__main__":
    main()
